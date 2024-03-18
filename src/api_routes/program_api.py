import io
from io import BytesIO
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, UploadFile, File, Request
from openpyxl.styles import Alignment, Font, Border, Side, Protection
from pydantic import BaseModel
from sqlalchemy import func

from src.db.session import session_scope
from src.helpers.utils import insert_exam_result, insert_course_work, get_student_from_uaa
from src.models import ExamCategory, ProgramSemester, ExamResultSummary, ExamResult, StudentCourseRegistration
from src.modules.academic_year.service import AcademicYearCrud
from src.modules.by_law.by_law_classes import BYLAW
from src.modules.by_law.by_law_files.by_law_2019 import ByLaw2019
from src.modules.by_law.service import ByLawCrud
from src.modules.program_course.service import ProgramCourseCrud
from src.modules.program_semester.service import ProgramSemesterCrud
from src.modules.programs.service import ProgramService
from src.modules.semester_registration.service import SemesterRegistrationService
from src.modules.student.service import StudentService
from src.shared.models import StudentPChangeModel
from src.shared.response import Response
from src.shared.response_code import ResponseCode


from pyinstrument import Profiler

program_router = APIRouter()
root_path = "/program"

from starlette.responses import StreamingResponse
from openpyxl import Workbook


@program_router.get(root_path)
async def get_program_data(code: str | None = None, uid: str | None = None):
    if code:
        return await ProgramService.api_get_program_by(code=code)
    elif uid:
        return await ProgramService.api_get_program_by(uid=uid)
    else:
        return Response(status=False, code=ResponseCode.NO_RECORD_FOUND,
                        message="Program Not Found", data=None)


class ProgramDepartmentInput(BaseModel):
    departments: List[str]


# class StudentPChangeInput(BaseModel):
#     students: List[str]


# @program_router.get("/program")
# async def get_program_data(parm: ProgramCodeInput):
#     print(parm)
#     return await ProgramService.api_get_program_by_code(parm)


@program_router.get("/programs")
async def get_program_data():
    return await ProgramService.api_get_programs()



@program_router.get("/profiling")
async def get_profiling(request: Request, call_next):
    # @program_router.middleware("http")
    profiling = True
    if profiling:
        profiler = Profiler(interval=settings.profiling_interval, async_mode="enabled")
        profiler.start()
        await call_next(request)
        profiler.stop()
        return HTMLResponse(profiler.output_html())
    else:
        return await call_next(request)

# These will get all programs uid by passed list of department
@program_router.post("/program/department")
async def get_program_data(parm: ProgramDepartmentInput):
    return ProgramService.api_get_program_by_departments(parm.departments)


@program_router.get("/active-by-law")
def get_active_by_law():
    try:
        result = ByLawCrud.get_active_by_law()
    except Exception as e:
        print(e)
        result = None
    return result


@program_router.get("/active-academic-year")
def get_active_by_academic_year():
    try:
        result = AcademicYearCrud.get_active_academic_year()
    except Exception as e:
        print(e)
        result = None
    return result


@program_router.get("/program-for-nhif")
async def api_get_program_name_duration(uid: str):
    return await ProgramService.api_get_program_name_duration(uid)


@program_router.get('/get-semester-registered-students')
def get_semester_registered_students():
    try:
        student_uids = SemesterRegistrationService.get_active_year_student_semester_registrations()
        if student_uids:
            return Response(status=True, code=ResponseCode.SUCCESS,
                            message="Success", data=student_uids)
    except Exception as e:
        print(e)
    return Response(status=False, code=ResponseCode.NO_RECORD_FOUND, message="No Record", data=[])


@program_router.get("/program-change-students")
async def api_get_program_change_student_list():
    return await ProgramService.api_get_program_change_student_list()


@program_router.post("/get-uqf-pchanges")
async def api_get_uqf_pchanges_list(parm: StudentPChangeModel):
    return await ProgramService.api_get_uqf_pchanges_list(parm)


@program_router.post("/generate-allocation-template/")
def generate_allocation_xls_template(allocation_uid: str, out_off: int, exam_category: int, assessment_number: int,
                                     assessment_weight: int):
    result = StudentService().get_allocation_students(allocation_uid)
    # Create a new workbook
    workbook = Workbook()

    # Create a new worksheet
    worksheet = workbook.active
    # Set column widths
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 45

    # Set the font style to Times New Roman
    font = Font(name="Times New Roman")
    font_border = Font(name="Times New Roman", bold=True)
    # Set the border style
    border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))

    # Define the vertical headers
    vertical_headers = ["Course Ante", "Program Code", "Academic Year", "Study Year", "Exam Category", "Assessment No",
                        "Mark Out of",
                        "Assessment Weight"]
    # Sample data for the vertical header
    data = {
        "Course Ante": result["program_course"].course.code,
        "Program Code": result["program_course"].program_semester.program.code,
        "Academic Year": result["program_course"].program_semester.academic_year.name,
        "Study Year": str(result["program_course"].program_semester.study_year),
        "Exam Category": str(exam_category),
        "Assessment No": str(assessment_number),
        "Mark Out of": str(out_off),
        "Assessment Weight": str(assessment_weight)
    }
    worksheet.sheet_view.showGridLines = False
    # Generate the data for the vertical header
    vertical_data = [data[header] for header in vertical_headers]
    for row, header in enumerate(vertical_headers, start=1):
        cell = worksheet.cell(row=row, column=1, value=header)
        cell.alignment = Alignment(horizontal='left')
        cell.font = font_border
        cell.border = None
        cell.protection = Protection(locked=False)
    for row, value in enumerate(vertical_data, start=1):
        cell = worksheet[f"C{row}"]
        cell.value = value
        cell.font = font_border
        cell.border = None
        cell.protection = Protection(locked=False)
    # Define the horizontal headers
    # worksheet.sheet_view.showGridLines = True
    horizontal_headers = ["SN", "Reg No", "Name", "Marks"]
    start_col = 1  # Start column for horizontal headers
    start_row = len(vertical_headers) + 1  # Start row for horizontal headers

    for col, header in enumerate(horizontal_headers, start=start_col):
        cell = worksheet.cell(row=start_row, column=col, value=header)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = font_border
        cell.border = border

    count = 0
    # print('data ===>', result['data'])
    for row, item in enumerate(result['data'], start=10):
        count += 1
        worksheet[f"A{row}"] = count
        worksheet[f"B{row}"] = item['registration_number']
        worksheet[f"C{row}"] = reformat_name(item['full_name'])
        worksheet[f"D{row}"] = ""
        # Align the cells to the center
        for col in range(1, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = font
            cell.border = border
    worksheet.cell(row=1, column=4, value=str(result["program_course"].id))
    # Set the specific column where cells should be non-editable (except column D)
    editable_column = 'D'

    # Iterate over rows in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            # Check if the current column is the editable column and cell row is not greater than 9
            if cell.column_letter == editable_column and cell.row >= 10:
                # Set protection to False for the editable column
                cell.protection = Protection(locked=False)
                cell.number_format = '0.00'
                # if cell.value is not None and float(int(cell.value)) > out_off:
                #     cell.value = 0  # Set the value to 10 if it's greater than 10
            else:
                # Set protection to True for other columns
                cell.protection = Protection(locked=True)

        # Protect the worksheet to make cells not editable
        worksheet.protection.sheet = True

    # Save the workbook
    # workbook.save("layout.xlsx")

    # Save the workbook to a BytesIO buffer
    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    # Set the appropriate headers for the response
    headers = {
        "Content-Disposition": f"attachment; filename=allocation_template.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    # Return the workbook as a streaming response
    return StreamingResponse(content=file_buffer, headers=headers)


def reformat_name(full_name: str) -> str:
    # Split the input string into parts
    parts = full_name.split()

    # Capitalize the first part
    parts[0] = parts[0].upper() + ', '

    # Capitalize the first letter of the second part
    parts[1] = parts[1][0].upper() + parts[1][1:].lower()

    # Capitalize the first letter of the third part
    if len(parts) > 2:
        parts[2] = parts[2][0].upper() + parts[2][1:].lower()

    # Join the parts back into a single string
    full_name_ = ' '.join(parts)

    return full_name_


@program_router.post("/generate-course-result-report/")
def generate_course_result_report(allocation_uid: str):
    result = StudentService().get_allocation_students(allocation_uid)
    # Create a new workbook
    workbook = Workbook()

    # Create a new worksheet
    worksheet = workbook.active
    # Set column widths
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 45

    # Set the font style to Times New Roman
    font = Font(name="Times New Roman")
    font_border = Font(name="Times New Roman", bold=True)
    # Set the border style
    border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))

    # Define the vertical headers
    vertical_headers = ["Course Ante", "Program Code", "Academic Year", "Study Year", "Exam Category", "Assessment No",
                        "Mark Out of",
                        "Assessment Weight"]
    # Sample data for the vertical header
    data = {
        "Course Ante": result["program_course"].course.code,
        "Program Code": result["program_course"].program_semester.program.code,
        "Academic Year": result["program_course"].program_semester.academic_year.name,
        "Study Year": str(result["program_course"].program_semester.study_year),
        "Exam Category": str(exam_category),
        "Assessment No": str(assessment_number),
        "Mark Out of": str(out_off),
        "Assessment Weight": str(assessment_weight)
    }
    worksheet.sheet_view.showGridLines = False
    # Generate the data for the vertical header
    vertical_data = [data[header] for header in vertical_headers]
    for row, header in enumerate(vertical_headers, start=1):
        cell = worksheet.cell(row=row, column=1, value=header)
        cell.alignment = Alignment(horizontal='left')
        cell.font = font_border
        cell.border = None
        cell.protection = Protection(locked=False)
    for row, value in enumerate(vertical_data, start=1):
        cell = worksheet[f"C{row}"]
        cell.value = value
        cell.font = font_border
        cell.border = None
        cell.protection = Protection(locked=False)
    # Define the horizontal headers
    # worksheet.sheet_view.showGridLines = True
    horizontal_headers = ["SN", "Name", "Reg No", "Sex"]
    start_col = 1  # Start column for horizontal headers
    start_row = len(vertical_headers) + 1  # Start row for horizontal headers

    for col, header in enumerate(horizontal_headers, start=start_col):
        cell = worksheet.cell(row=start_row, column=col, value=header)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = font_border
        cell.border = border

    count = 0
    for row, item in enumerate(result['data'], start=10):
        count += 1
        worksheet[f"A{row}"] = count
        worksheet[f"B{row}"] = item['full_name']
        worksheet[f"C{row}"] = item['registration_number']
        worksheet[f"D{row}"] = item['gender']
        # Align the cells to the center
        for col in range(1, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = font
            cell.border = border
    worksheet.cell(row=1, column=4, value=str(result["program_course"].id))
    # Set the specific column where cells should be non-editable (except column D)
    editable_column = 'D'

    # Iterate over rows in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:

            # lock all cells of the excel sheet
            cell.protection = Protection(locked=True)
            # Check if the current column is the editable column and cell row is not greater than 9
            # if cell.column_letter == editable_column and cell.row >= 10:
            #     # Set protection to False for the editable column
            #     cell.protection = Protection(locked=False)
            #     cell.number_format = '0.00'
            #     # if cell.value is not None and float(int(cell.value)) > out_off:
            #     #     cell.value = 0  # Set the value to 10 if it's greater than 10
            # else:
            #     # Set protection to True for other columns
            #     cell.protection = Protection(locked=True)

            # Construct the query
            query = (
                session.query(
                    ProgramCourses.id.label("program_course_id"),
                    Course.code.label("course_code"),
                    Course.name.label("course_name"),
                    ExamCoursework.score.label("score"),
                    ProgramCourseAssessments.minimum_exams.label("minimum_exams"),
                    ExamCategory.name.label("assessment_type")
                )
                .join(ProgramCourseAssessments, ProgramCourses.id == ProgramCourseAssessments.program_course_id)
                .join(Course, ProgramCourses.course_id == Course.id)
                .join(ExamCategory, ProgramCourseAssessments.exam_category_id == ExamCategory.id)
                .join(ExamCoursework, ExamCoursework.program_course_id == ProgramCourses.id)
                .filter(
                    ExamCoursework.student_uid.in_(student_ids),
                    ProgramCourseAssessments.minimum_exams == 2  # Filter by minimum exams
                )
                .order_by(ProgramCourseAssessments.id,
                          ExamCategory.id)  # Order by program course assessments and assessment type
                .all()
            )

            # for row_ in query:
            #     print(row_)

        # Protect the worksheet to make cells not editable
        worksheet.protection.sheet = True

    # Save the workbook
    # workbook.save("layout.xlsx")

    # Save the workbook to a BytesIO buffer
    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    # Set the appropriate headers for the response
    headers = {
        "Content-Disposition": f"attachment; filename=allocation_template.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    # Return the workbook as a streaming response
    return StreamingResponse(content=file_buffer, headers=headers)


@program_router.post("/extract-data")
async def extract_data(file: UploadFile = File(...)):
    contents = await file.read()

    # Load the workbook from the file contents
    workbook = openpyxl.load_workbook(io.BytesIO(contents))

    # Get the desired worksheet by name or index
    worksheet = workbook.active  # Modify this line with the appropriate worksheet name or index
    exam_category_id = worksheet.cell(row=5, column=3).value
    assessment_number = worksheet.cell(row=6, column=3).value
    out_off = float(worksheet.cell(row=7, column=3).value)
    program_course_id = worksheet.cell(row=1, column=4).value
    weight = float(worksheet.cell(row=8, column=3).value)

    # Assuming the data is in a specific sheet and columns
    sn_column = 1  # Assuming SN is in column A
    reg_no_column = 2  # Assuming Reg No is in column B
    name_column = 3  # Assuming Name is in column C
    marks_column = 4  # Assuming Marks is in column D

    with session_scope() as session:
        is_ue = session.query(ExamCategory).filter(
            ExamCategory.id == exam_category_id).first().exam_category_group.is_ue
        # get student list from uaa service to get student uid after filtering
        students = get_student_from_uaa()
        success = 0
        failed = 0
        failed_students = []
        response_data = {
            "success": 0,
            "failed": 0,
            "failed_students": []
        }
        for row in worksheet.iter_rows(min_row=10, values_only=True):
            reg_number = row[reg_no_column - 1]
            score = float(row[marks_column - 1])
            # Find the item with the specified registration_number
            if students:
                matching_item = next(
                    (item for item in students if item["registration_number"] == reg_number), None)
                if matching_item:
                    student_uid = matching_item["uid"]
                    # print('student',reg_number,student_uid)
                    if score <= out_off:
                        if is_ue:
                            result = insert_exam_result(student_uid, program_course_id, exam_category_id, score,
                                                        out_off,
                                                        weight)
                            if result:
                                success = success + 1
                            else:
                                failed = failed + 1
                                failed_students.append({"reg_number": reg_number, "reason": "Data processing error"})
                        else:
                            result = insert_course_work(student_uid, program_course_id, exam_category_id,
                                                        assessment_number,
                                                        out_off, score,
                                                        weight)
                            if result:
                                if result:
                                    success = success + 1
                                else:
                                    failed = failed + 1
                                    failed_students.append(
                                        {"reg_number": reg_number, "reason": "Data processing error"})

                    else:
                        failed = failed + 1
                        failed_students.append({"reg_number": reg_number, "reason": "Score is greater than out off"})
                else:
                    failed = failed + 1
                    failed_students.append(
                        {"reg_number": reg_number, "reason": "Data processing error ,student not found"})

            else:
                failed = failed + 1
                failed_students.append(
                    {"reg_number": reg_number, "reason": "Data processing error , UAA service not found"})

    # Save the extracted data in the database
    # ...
    response_data["failed_students"] = failed_students
    response_data["failed"] = failed
    response_data["success"] = success
    return response_data


@program_router.post("/generate-exam-semester-results/")
def generate_exam_semester_results(program_uid: str, academic_year_uid: str, semester: int, year_of_study: int):
    with (session_scope() as session):
        # result = StudentService().get_allocation_students(allocation_uid)
        # Create a new workbook
        workbook = Workbook()
        program = ProgramService.get_program_by_uid(program_uid)
        if program is None:
            return {
                "status": False,
                "message": "Program Is Not Found"
            }

        academic_year = AcademicYearCrud.get_academic_year_by_uid(academic_year_uid)
        if academic_year is None:
            return {
                "status": False,
                "message": "Academic Year Is Not Found"
            }

        # Get Program Semester Information
        program_semester = ProgramSemesterCrud.get_program_semester_by_data(semester=semester,
                                                                            year_of_study=year_of_study,
                                                                            program_id=program.id,
                                                                            academic_year_id=academic_year.id)
        if not program_semester:
            return {
                "status": False,
                "message": "Program Semester Is Not Defined"
            }
        # Create a new worksheet
        worksheet = workbook.active
        # Set column widths
        # worksheet.column_dimensions['A'].width = 15
        # worksheet.column_dimensions['B'].width = 20
        # worksheet.column_dimensions['C'].width = 45

        # Set the font style to Times New Roman
        font = Font(name="Times New Roman", size=12)
        font_border = Font(name="Times New Roman", bold=True, size=12)
        # Set the border style
        border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))

        # Merge cells from C1 to J1
        worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
        # Merge cells from C2 to J2
        worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=10)

        merged_cell_h1 = worksheet.cell(row=1, column=3, value="SOKOINE UNIVERSITY OF AGRICULTURE")
        merged_cell_h1.alignment = Alignment(horizontal='center')
        merged_cell_h1.font = font
        # Merge cells from C2 to J2
        merged_cell_h2 = worksheet.cell(row=2, column=3, value="P.O. BOX 3000, CHUO KIKUU, MOROGORO, TANZANIA")
        merged_cell_h2.alignment = Alignment(horizontal='center')
        merged_cell_h2.font = font
        count_rows = 4

        # Define the vertical headers
        vertical_headers = ["Semester Result", "Program", "Program Code", "Academic Year", "Study Year", "Semester",
                            "Semester Credit"]
        # Sample data for the vertical header
        data = {
            "Semester Result": "",
            "Program": str(program.name),
            "Program Code": str(program.code),
            "Academic Year": str(academic_year.name),
            "Study Year": str(year_of_study),
            "Semester": str(semester),
            "Semester Credit": str(program_semester.core_credits + program_semester.elective_credits)
        }
        worksheet.sheet_view.showGridLines = False

        # Generate the data for the vertical header
        vertical_data = [data[header] for header in vertical_headers]
        for row, header in enumerate(vertical_headers, start=4):
            count_rows += 1
            cell = worksheet.cell(row=row, column=3)
            cell.value = header
            cell.alignment = Alignment(horizontal='left')
            cell.font = font_border
            cell.border = None
            cell.protection = Protection(locked=False)
        for row, value in enumerate(vertical_data, start=4):
            cell = worksheet[f"F{row}"]
            cell.value = value
            cell.font = font_border
            cell.border = None
            cell.protection = Protection(locked=False)

        # Define the horizontal headers
        # worksheet.sheet_view.showGridLines = True
        start_col = 1  # Start column for horizontal headers

        worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=16)
        summary_text = worksheet.cell(row=count_rows, column=1, value="Summary")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border
        count_rows += 1
        worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=2)
        summary_text = worksheet.cell(row=count_rows, column=1, value="Status")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border

        summary_text = worksheet.cell(row=count_rows, column=3, value="Male")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border

        summary_text = worksheet.cell(row=count_rows, column=4, value="Female")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border

        worksheet.merge_cells(start_row=count_rows, start_column=5, end_row=count_rows, end_column=7)
        summary_text = worksheet.cell(row=count_rows, column=5, value="Total")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border

        worksheet.merge_cells(start_row=count_rows, start_column=8, end_row=count_rows, end_column=10)
        summary_text = worksheet.cell(row=count_rows, column=8, value="Male %")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border

        worksheet.merge_cells(start_row=count_rows, start_column=11, end_row=count_rows, end_column=13)
        summary_text = worksheet.cell(row=count_rows, column=11, value="Female %")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border

        worksheet.merge_cells(start_row=count_rows, start_column=14, end_row=count_rows, end_column=16)
        summary_text = worksheet.cell(row=count_rows, column=14, value="Total %")
        summary_text.alignment = Alignment(horizontal='center')
        summary_text.font = font_border
        summary_text.border = border
        count_rows += 1

        status_info = ["CONTINUING", "PROBATION", "INCOMPLETE", "POSTPONE", "RETAKE"]
        status_data = []
        for status in status_info:
            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=2)
            cell = worksheet.cell(row=count_rows, column=1, value=status)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = font_border
            cell.border = border
            count_rows += 1

        # skip status

        # return status
        # Get Courses
        total_elective_courses = 0
        total_core_courses = 0
        program_courses = ProgramCourseCrud.get_program_course_by_program_semester(
            program_semester_id=program_semester.id)
        print("program_courses:===============>", program_courses)
        if program_courses is not None:
            courses_list = ""
            for pc in program_courses:
                if pc.course_category.name.upper() == 'CORE':
                    total_core_courses += 1
                elif pc.course_category.name.upper() == 'ELECTIVE':
                    total_elective_courses += 1
                courses_list += f"{pc.course.code}: {pc.course.name}, "

            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=16)
            cell = worksheet.cell(row=count_rows, column=1, value=courses_list)
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True, indent=1)
            cell.font = font
            cell.border = border
            # Calculate the number of lines in the wrapped text
            text_length = len(courses_list)
            num_lines = max(1, text_length // 100)  # Assuming an average line width of 100 characters, adjust as needed
            # Set the row height based on the number of lines
            row_height_per_line = 10  # Adjust as needed
            worksheet.row_dimensions[count_rows].height = num_lines * row_height_per_line
            count_rows += 1

            course_colum_merge = 5
            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows,
                                  end_column=course_colum_merge)
            summary_text = worksheet.cell(row=count_rows, column=1, value="Course")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            # Core Course Title
            worksheet.merge_cells(start_row=count_rows, start_column=course_colum_merge + 1, end_row=count_rows,
                                  end_column=course_colum_merge + total_core_courses)
            summary_text = worksheet.cell(row=count_rows, column=course_colum_merge + 1, value="Core")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border
            course_colum_merge += total_core_courses
            # course_colum_merge += 1

            # print("S:", total_core_courses)
            # print("Sum of Elective Courses:", total_elective_courses)
            # Elective Course Title
            worksheet.merge_cells(start_row=count_rows, start_column=course_colum_merge + 1, end_row=count_rows,
                                  end_column=total_elective_courses + course_colum_merge)
            summary_text = worksheet.cell(row=count_rows, column=course_colum_merge + 1, value="Elective")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border
            count_rows += 1

            credit_hr_colum_merge = 5
            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows,
                                  end_column=credit_hr_colum_merge)
            summary_text = worksheet.cell(row=count_rows, column=1, value="Credit Hours")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            for pc in program_courses:
                credit_hr_colum_merge += 1
                cel = worksheet.cell(row=count_rows, column=credit_hr_colum_merge, value=pc.credit)
                cel.alignment = Alignment(horizontal='center')
                cel.font = font_border
                cel.border = border

            count_rows += 1

            # SN
            summary_text = worksheet.cell(row=count_rows, column=1, value="SNo")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border
            # Name
            text = worksheet.cell(row=count_rows, column=2, value="Name")
            text.alignment = Alignment(horizontal='center')
            text.font = font_border
            text.border = border

            # SN
            text = worksheet.cell(row=count_rows, column=3, value="Reg. No.")
            text.alignment = Alignment(horizontal='center')
            text.font = font_border
            text.border = border

            text = worksheet.cell(row=count_rows, column=4, value="Sex")
            text.alignment = Alignment(horizontal='center')
            text.font = font_border
            text.border = border

            text = worksheet.cell(row=count_rows, column=5, value="Examination No.")
            text.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, indent=1)
            text.font = font_border
            text.border = border

            course_code_colum_merge = 5
            for pc in program_courses:
                course_code_colum_merge += 1
                cel = worksheet.cell(row=count_rows, column=course_code_colum_merge, value=pc.course.code)
                cel.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
                cel.font = font_border
                cel.border = border
                # Set the row height based on the number of lines
                row_height_per_line = 10  # Adjust as needed
                worksheet.row_dimensions[count_rows].height = 8 * row_height_per_line
            total_titles = ["TOTAL CREDITS TAKEN", "CREDITS ACQUIRED", "CORE COURSES FAILED", "GPA", "Remarks",
                            "Courses Under Probation"]
            for title in total_titles:
                course_code_colum_merge += 1
                start_column = course_code_colum_merge
                end_column = course_code_colum_merge

                # Set the value of the cell before merging
                cel = worksheet.cell(row=count_rows - 2, column=start_column, value=title)
                if title == 'Remarks' or title == 'Courses Under Probation':
                    cel.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, indent=1)
                else:
                    cel.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
                    # Increment the course_code_colum_merge by the width of the merged area minus 1
                    course_code_colum_merge += (end_column - start_column)
                    # Set the row height based on the number of lines
                    row_height_per_line = 10  # Adjust as needed
                    worksheet.row_dimensions[count_rows].height = 12 * row_height_per_line

                cel.font = font_border
                cel.border = border
                # Merge cells after setting the value
                worksheet.merge_cells(start_row=count_rows - 2, start_column=start_column, end_row=count_rows,
                                      end_column=end_column)
            # Get students list
            results = session.query(ExamResultSummary.student_uid, func.max(ExamResultSummary.gender).label('sex'),
                                    func.max(ExamResultSummary.registration_number).label('registration_number'),
                                    func.concat(func.max(ExamResultSummary.first_name), ' ',
                                                func.max(ExamResultSummary.middle_name), ' ',
                                                func.max(ExamResultSummary.last_name)).label('full_name')).filter(
                ExamResultSummary.program_uid == program_uid, ExamResultSummary.academic_year_uid == academic_year_uid,
                ExamResultSummary.semester == semester, ExamResultSummary.study_year == year_of_study).group_by(
                ExamResultSummary.student_uid).all()
            if len(results) > 0:
                count_rows += 1
                count = 0
                for row, item in enumerate(results, start=count_rows):
                    count += 1
                    total_credit_hrs_taken = 0
                    total_credit_hrs_acquired = 0
                    # count_rows += 1
                    worksheet[f"A{row}"] = count
                    worksheet[f"B{row}"] = reformat_name(item['full_name'])
                    worksheet[f"C{row}"] = item['registration_number']
                    worksheet[f"D{row}"] = item['sex'][0]
                    worksheet[f"E{row}"] = "---"
                    for col in range(1, 5):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.font = font
                        cell.border = border
                    col = 5
                    for pc in program_courses:
                        col += 1
                        exam_result = session.query(ExamResultSummary.grade, ExamResultSummary.grade_remark).filter(
                            ExamResultSummary.student_uid == item['student_uid'],
                            ExamResultSummary.program_course_id == pc.id).first()
                        if exam_result:
                            value = exam_result.grade
                            # Check if student have passed this course sum
                            if exam_result.grade_remark.upper() == "PASS":
                                total_credit_hrs_acquired += pc.credit
                        else:
                            value = '-'
                        cel = worksheet.cell(row=count_rows, column=col, value=value)
                        cel.alignment = Alignment(horizontal='center')
                        cel.font = font
                        cel.border = border

                        # Check if student have registered this course sum its credit hour to total_credit_hrs_taken
                        if session.query(StudentCourseRegistration.id).filter(
                                StudentCourseRegistration.student_uid == item['student_uid'],
                                StudentCourseRegistration.program_course_id == pc.id).first():
                            total_credit_hrs_taken += pc.credit
                    total_title_results = [total_credit_hrs_taken, total_credit_hrs_acquired, "-",
                                           "-", "-", "-"]
                    for title in total_title_results:
                        col += 1
                        cel = worksheet.cell(row=count_rows, column=col, value=title)
                        cel.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True,
                                                  indent=1)
                        cel.font = font
                        cel.border = border
                    count_rows += 1
                worksheet.column_dimensions['B'].width = 30
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 3

        count = 0
        # print('data ===>', result['data'])
        # for row, item in enumerate(result['data'], start=10):
        #     count += 1
        #     worksheet[f"A{row}"] = count
        #     worksheet[f"B{row}"] = item['registration_number']
        #     worksheet[f"C{row}"] = reformat_name(item['full_name'])
        #     worksheet[f"D{row}"] = ""
        #     # Align the cells to the center
        #     for col in range(1, 5):
        #         cell = worksheet.cell(row=row, column=col)
        #         cell.alignment = Alignment(horizontal='left', vertical='center')
        #         cell.font = font
        #         cell.border = border
        # worksheet.cell(row=1, column=4, value=str(result["program_course"].id))
        # Set the specific column where cells should be non-editable (except column D)
        editable_column = 'D'

        # Iterate over rows in the worksheet
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.protection = Protection(locked=True)

            # Protect the worksheet to make cells not editable
            worksheet.protection.sheet = True

        # Save the workbook
        # workbook.save("layout.xlsx")

        # Save the workbook to a BytesIO buffer
        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        # Set the appropriate headers for the response
        headers = {
            "Content-Disposition": f"attachment; filename=allocation_template.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

        # Return the workbook as a streaming response
        return StreamingResponse(content=file_buffer, headers=headers)
