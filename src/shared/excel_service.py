import io
import base64
import logging
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.shared.excel_types import Base64ExcelOutput

logger = logging.getLogger(__name__)


# ─── Styles ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="medium"),
)

# ─── Column definition helper ─────────────────────────────────────────────────
class TemplateColumn:
    """Describes a single column in an Excel import template."""

    def __init__(self, header: str, width: int = 25, example: Optional[str] = None):
        self.header = header
        self.width = width
        self.example = example


# ─── Service ──────────────────────────────────────────────────────────────────
class ExcelTemplateService:

    @staticmethod
    def generate_template(
        columns: List[TemplateColumn],
        sheet_name: str = "Template",
        file_name: str = "import_template.xlsx",
    ) -> Base64ExcelOutput:
        """
        Generate an Excel template workbook, style headers, and return
        it as a base64-encoded output.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # ── Write header row ──────────────────────────────────────────────
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

        # ── Column widths ─────────────────────────────────────────────────
        for col_idx, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col.width

        # ── Optional example row ──────────────────────────────────────────
        if any(c.example for c in columns):
            for col_idx, col in enumerate(columns, start=1):
                if col.example is not None:
                    ws.cell(row=2, column=col_idx, value=col.example)

        # ── Freeze top row ────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        raw_bytes = buffer.read()

        b64_data = base64.b64encode(raw_bytes).decode("utf-8")

        return Base64ExcelOutput(file_name=file_name, base64_data=b64_data)

    # ──────────────────────────────────────────────────────────────────────────
    @staticmethod
    def parse_upload(base64_data: str) -> List[dict]:
        """
        Decode a base64 Excel file and return a list of dictionaries
        (one per data row, keyed by the header text).
        """
        try:
            padding = 4 - len(base64_data) % 4
            if padding != 4:
                base64_data += "=" * padding

            raw = base64.b64decode(base64_data)
            wb = openpyxl.load_workbook(io.BytesIO(raw))
            ws = wb.active

            # Read headers from the first row
            headers: List[str] = []
            for cell in ws[1]:
                if cell.value is None:
                    break
                headers.append(str(cell.value).strip())

            # Read data rows
            rows: List[dict] = []
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row_cells):
                    continue  # skip fully-empty rows
                row_dict: dict = {}
                for idx, header in enumerate(headers):
                    val = row_cells[idx] if idx < len(row_cells) else None
                    if isinstance(val, str):
                        val = val.strip()
                    row_dict[header] = val
                rows.append(row_dict)

            return rows

        except Exception as exc:
            logger.exception("Failed to parse uploaded Excel data")
            raise ValueError(f"Invalid or corrupted Excel file: {exc}")