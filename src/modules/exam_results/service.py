import base64
import json
import math
from io import BytesIO
from typing import List

import pendulum
import requests
from openpyxl.workbook import Workbook
from sqlalchemy import select, func

from src.core.config import settings
from src.db.session import session_scope
from src.helpers.utils import to_superscript, get_user_unit_department_headship
from src.models import ExamResult, ExamResultSummary, ProgramSemester, ProgramCourse, Program, AcademicYear
from src.models.exam_course_result_forward_logs import ExamCourseResultForwardLogs
from src.modules.academic_year.service import AcademicYearService
from src.modules.programs.service import ProgramService
from src.shared.response import Response
from src.shared.response_code import ResponseCode
from src.types import ExamResultInput, ExamResultNode, ExcelFile
import openpyxl
from fastapi import APIRouter, UploadFile, File
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, Protection, PatternFill
from openpyxl.utils import get_column_letter


class ExamResultService:
    @staticmethod
    def get_exam_results() -> List[ExamResult]:
        with session_scope() as session:
            result = session.query(ExamResult).filter(ExamResult.deleted_at.is_(None)).all()
            return result

    @staticmethod
    def generate_partial_transcript(registration_number, info) -> Response[ExcelFile]:
        with (session_scope() as session):
            workbook = Workbook()
            if info.context.user is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Session Expired, refresh this page",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            headers = {
                "Content-Type": "application/json"
            }

            data_obj = {
                "search": registration_number,
                "offset": "0",
                "limit": "1"
            }
            payload = json.dumps(data_obj)
            # Get student data
            get_student_response = requests.post(
                settings.UAA_URi + f'/students/search', data=payload,
                headers=headers)

            # response = requests.post(settings.UAA_URi + f'/students-details-by-uids', data=payload,
            #                          headers=headers)
            get_student_response.raise_for_status()
            if get_student_response.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Get Student Data",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            student_data = get_student_response.json()
            if not student_data:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Please upload your signature before downloading Semester results",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            student = student_data[0]
            program_uid = student['programme_uid']
            student_full_name = student['user']['full_name']
            first_name = student['user']['first_name']
            middle_name = student['user']['middle_name']
            last_name = student['user']['last_name']
            date_of_birth = student['user']['date_of_birth']
            gender = student['user']['gender']
            campus = student['user']['campus']
            country = student['user']['country']
            student_uid = student['uid']
            first_name = student['user']['first_name']
            registration_number = student['registration_number']
            user_uid = student['user']['uid']
            # print(student)
            # Send the POST request
            response = requests.post(
                settings.UAA_URi + f'/user/document?user_uid={info.context.user.uid}&document_type=signature',
                headers=headers)

            response.raise_for_status()
            if response.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Get User Signature",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            signature_data = response.json()
            if not signature_data['status']:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Please upload your signature before downloading Semester results",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            # Get student profile picture
            response = requests.post(
                settings.UAA_URi + f'/user/document?user_uid={user_uid}&document_type=profile',
                headers=headers)

            response.raise_for_status()
            if response.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Get Student Profile Picture",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            profile_picture_data = response.json()
            if not profile_picture_data['status']:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Student Have not Uploaded Profile Picture",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            program = ProgramService.get_program_by_uid(program_uid)
            get_headship = requests.get(
                settings.UAA_URi + f'/user/check-user-headship?user_uid={info.context.user.uid}&'
                                   f'department_uid={program.department_uid}&headship_type=unit', headers=headers)
            get_headship.raise_for_status()
            if get_headship.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Find Principal Headship",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            headship = get_headship.json()
            if headship['status'] != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message=headship['message'],
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            file_name = f"{student_full_name} ({registration_number}) PARTIAL TRANSCRIPT"
            # Create a new worksheet
            worksheet = workbook.active
            # Set the orientation to landscape
            worksheet.print_options.horizontalCentered = True
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

            # Set the font style to Times New Roman
            font = Font(name="Times New Roman", size=12)
            small_font = Font(name="Times New Roman", size=10)
            small_font_p = Font(size=small_font.size * 0.75)
            fill_color = PatternFill(start_color='FF999999', end_color='FF999999', fill_type='solid')
            font_border = Font(name="Times New Roman", bold=True, size=12)
            small_font_border = Font(name="Times New Roman", bold=True, size=9)
            # Set the border style
            border = Border(left=Side(border_style="thin"),
                            right=Side(border_style="thin"),
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))

            # Merge cells from C1 to J1
            worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
            # Merge cells from C2 to J2
            worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=10)

            merged_cell_h1 = worksheet.cell(row=1, column=3, value="SOKOINE UNIVERSITY OF AGRICULTURE")
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = font
            # Merge cells from C2 to J2
            merged_cell_h2 = worksheet.cell(row=2, column=3, value="P.O. BOX 3000, CHUO KIKUU, MOROGORO, TANZANIA")
            merged_cell_h2.alignment = Alignment(horizontal='center')
            merged_cell_h2.font = font
            count_rows = 3
            worksheet.merge_cells(start_row=count_rows, start_column=3, end_row=count_rows, end_column=10)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=3, value="PARTIAL TRANSCRIPT")
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = font
            count_rows += 1
            worksheet.merge_cells(start_row=count_rows, start_column=3, end_row=count_rows, end_column=10)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=3, value=campus)
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = font
            count_rows += 1
            worksheet.merge_cells(start_row=count_rows, start_column=3, end_row=count_rows, end_column=10)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=3, value=program.name)
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = small_font

            count_rows = 7
            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows + 1, end_column=2)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=1, value="STUDENT INFORMATION")
            merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
            merged_cell_h1.font = font_border
            merged_cell_h1.border = border

            image_path = "assets/logo/logo.png"

            # Read the image
            with open(image_path, "rb") as f:
                image_data = f.read()

            # Encode the image data as base64
            image_base64 = base64.b64encode(image_data).decode()

            decoded_image = base64.b64decode(image_base64)
            # Load the image into BytesIO
            image_stream = BytesIO(decoded_image)

            # Create an Image object from BytesIO
            img = Image(image_stream)
            # Resize the image (adjust the width and height as needed)
            img.width = 150
            img.height = 100

            # Calculate the column letter for the anchor cell (e.g., "B" for column number 2)
            anchor_column = openpyxl.utils.get_column_letter(1)
            anchor_row = 1
            # Set the row number for the anchor cell

            # # Set the anchor cell for the image to the bottom center cell
            anchor_cell = f'{anchor_column}{anchor_row}'
            # print(img)
            # Add image to the worksheet and set the anchor cell
            worksheet.add_image(img, anchor_cell)

            # Decode the Base64 image
            base64_image = profile_picture_data['data']['base64doc']
            decoded_image = base64.b64decode(base64_image)

            # Load the image into BytesIO
            image_stream = BytesIO(decoded_image)

            # Create an Image object from BytesIO
            img = Image(image_stream)
            # Resize the image (adjust the width and height as needed)
            img.width = 150
            img.height = 100

            # Calculate the column letter for the anchor cell (e.g., "B" for column number 2)
            # Calculate the column letter for the anchor cell (e.g., "B" for column number 2)
            anchor_column = openpyxl.utils.get_column_letter(11)
            anchor_row = 1

            # # Set the anchor cell for the image to the bottom center cell
            anchor_cell = f'{anchor_column}{anchor_row}'
            # Add image to the worksheet and set the anchor cell
            worksheet.add_image(img, anchor_cell)

            names_headers = ["SURNAME", "FIRST NAME", "OTHER NAME", "DATE OF BIRTH"]
            start_header_column = 3
            end_header_column = start_header_column + 1
            for name in names_headers:
                worksheet.merge_cells(start_row=count_rows, start_column=start_header_column, end_row=count_rows,
                                      end_column=end_header_column)
                merged_cell_h1 = worksheet.cell(row=count_rows, column=start_header_column, value=name)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border
                merged_cell_h1.border = border
                start_header_column += 2
                end_header_column += 2

            other_header_columns = ["GENDER", "NATIONALITY"]
            start_header_column -= 1
            for other_header in other_header_columns:
                merged_cell_h1 = worksheet.cell(row=count_rows, column=start_header_column + 1, value=other_header)
                merged_cell_h1.alignment = Alignment(horizontal='center')
                merged_cell_h1.border = border
                merged_cell_h1.font = small_font_border
                start_header_column += 1

            count_rows += 1
            header_column_data = [last_name, first_name, middle_name, date_of_birth, gender, country]
            start_header_column = 3
            end_header_column = start_header_column + 1
            no = 0
            for name in header_column_data:
                no += 1
                if no <= 4:
                    worksheet.merge_cells(start_row=count_rows, start_column=start_header_column, end_row=count_rows,
                                          end_column=end_header_column)
                if no == 6:
                    start_header_column -= 1
                merged_cell_h1 = worksheet.cell(row=count_rows, column=start_header_column, value=name)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font
                merged_cell_h1.border = border
                start_header_column += 2
                end_header_column += 2

            count_rows += 1

            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=3)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=1, value="REGISTRATION")
            merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
            merged_cell_h1.font = small_font_border
            merged_cell_h1.border = border
            # Value
            worksheet.merge_cells(start_row=count_rows + 1, start_column=1, end_row=count_rows + 1, end_column=3)
            merged_cell_h1 = worksheet.cell(row=count_rows + 1, column=1, value=registration_number)
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = small_font
            merged_cell_h1.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=4, end_row=count_rows, end_column=7)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=4, value="DATE OF ADMISSION")
            merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
            merged_cell_h1.font = small_font_border
            merged_cell_h1.border = border
            # Value
            worksheet.merge_cells(start_row=count_rows + 1, start_column=4, end_row=count_rows + 1, end_column=7)
            merged_cell_h1 = worksheet.cell(row=count_rows + 1, column=4, value='N/A')
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = small_font
            merged_cell_h1.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=8, end_row=count_rows, end_column=11)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=8, value="DATE OF COMPLETION")
            merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
            merged_cell_h1.font = small_font_border
            merged_cell_h1.border = border
            # Value
            worksheet.merge_cells(start_row=count_rows + 1, start_column=8, end_row=count_rows + 1, end_column=11)
            merged_cell_h1 = worksheet.cell(row=count_rows + 1, column=8, value='N/A')
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = small_font
            merged_cell_h1.border = border
            #
            merged_cell_h1 = worksheet.cell(row=count_rows, column=12, value='')
            merged_cell_h1.border = border
            merged_cell_h1 = worksheet.cell(row=count_rows + 1, column=12, value='')
            merged_cell_h1.border = border

            count_rows += 2

            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=3)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=1, value="TYPE OF SPONSOR")
            merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
            merged_cell_h1.font = small_font_border
            merged_cell_h1.border = border
            # Value
            worksheet.merge_cells(start_row=count_rows + 1, start_column=1, end_row=count_rows + 1, end_column=3)
            merged_cell_h1 = worksheet.cell(row=count_rows + 1, column=1, value='')
            merged_cell_h1.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=4, end_row=count_rows, end_column=12)
            merged_cell_h1 = worksheet.cell(row=count_rows, column=4, value="NAME OF SPONSOR")
            merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
            merged_cell_h1.font = small_font_border
            merged_cell_h1.border = border
            # Value
            worksheet.merge_cells(start_row=count_rows + 1, start_column=4, end_row=count_rows + 1, end_column=12)
            merged_cell_h1 = worksheet.cell(row=count_rows + 1, column=4, value='PRIVATE Sponsored student')
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = small_font
            merged_cell_h1.border = border

            count_rows += 1

            exam_result = session.query(ExamResultSummary.semester).filter(
                ExamResultSummary.registration_number == registration_number,
                ExamResultSummary.deleted_at.is_(None)). \
                group_by(ExamResultSummary.semester). \
                order_by(ExamResultSummary.semester.asc()).all()

            if exam_result:
                sum_grade_point_credit = 0
                total_credit_hrs_taken = 0
                total_credit_hrs_acquired = 0
                has_incomplete_course = False
                total_failed_core_subject = 0
                total_pass_core_subject = 0

                for exam in exam_result:
                    worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=12)
                    merged_cell_h1 = worksheet.cell(row=count_rows, column=1, value=f"SEMESTER {str(exam.semester)}")
                    merged_cell_h1.alignment = Alignment(horizontal='left', wrapText=True)
                    merged_cell_h1.font = font_border
                    merged_cell_h1.border = border
                    count_rows += 1
                    exam_result_header = ["S/NO", "Course Ante", "Course Description", "TYPE", "CREDIT", "GRADE"]
                    rows = 0

                    start_header_columns = 0
                    end_header_columns = 0
                    for e_header in exam_result_header:
                        rows += 1
                        if rows in [3]:
                            start_header_columns += 2
                            end_header_columns += 6
                            # print(f"{start_header_columns} : {end_header_columns}")
                            worksheet.merge_cells(start_row=count_rows, start_column=start_header_columns,
                                                  end_row=count_rows,
                                                  end_column=end_header_columns)
                            start_header_columns = 9

                        else:
                            start_header_columns += 1
                            if rows == 2:
                                end_header_columns += 2
                                # print(f"{start_header_columns} : {end_header_columns}")
                                worksheet.merge_cells(start_row=count_rows, start_column=start_header_columns,
                                                      end_row=count_rows,
                                                      end_column=end_header_columns)

                            else:
                                end_header_columns += 1

                        if rows == 3:
                            merged_cell_h1 = worksheet.cell(row=count_rows, column=4, value=e_header)
                            merged_cell_h1.alignment = Alignment(horizontal='center')
                        else:
                            merged_cell_h1 = worksheet.cell(row=count_rows, column=start_header_columns, value=e_header)
                            merged_cell_h1.alignment = Alignment(horizontal='center')
                        merged_cell_h1.font = small_font_border
                        merged_cell_h1.border = border
                    count_rows += 1
                    all_exam_result = session.query(
                        ExamResultSummary.grade_point_credit,
                        ExamResultSummary.grade,
                        ExamResultSummary.grade_remark,
                        ExamResultSummary.credit,
                        ExamResultSummary.course_code,
                        ExamResultSummary.course_name,
                        func.upper(ExamResultSummary.course_category).label('course_category'),
                        ExamResultSummary.credit
                    ).filter(
                        ExamResultSummary.student_uid == student_uid,
                        ExamResultSummary.semester == exam.semester,
                        ExamResultSummary.deleted_at.is_(None)). \
                        order_by(ExamResultSummary.course_category.asc()).all()
                    if all_exam_result:
                        no = 0
                        for result in all_exam_result:
                            value = result.grade
                            if result.grade_remark.upper() == "PASS":
                                total_pass_core_subject += 1
                            elif result.course_category.name.upper() == 'CORE' and value != 'I':
                                total_failed_core_subject += 1
                            else:
                                has_incomplete_course = True

                            sum_grade_point_credit += result.grade_point_credit
                            total_credit_hrs_taken += result.credit
                            no += 1
                            exam_result_headers = [no, result.course_code, result.course_name, result.course_category,
                                                   result.credit, result.grade]
                            rows = 0
                            start_header_columns = 0
                            end_header_columns = 0
                            for e_header in exam_result_headers:
                                rows += 1
                                if rows in [3]:
                                    start_header_columns += 2
                                    end_header_columns += 6
                                    # print(f"{start_header_columns} : {end_header_columns}")
                                    worksheet.merge_cells(start_row=count_rows, start_column=start_header_columns,
                                                          end_row=count_rows,
                                                          end_column=end_header_columns)
                                    start_header_columns = 9
                                else:
                                    start_header_columns += 1
                                    if rows == 2:
                                        end_header_columns += 2
                                        # print(f"{start_header_columns} : {end_header_columns}")
                                        worksheet.merge_cells(start_row=count_rows, start_column=start_header_columns,
                                                              end_row=count_rows,
                                                              end_column=end_header_columns)

                                    else:
                                        end_header_columns += 1

                                if rows == 3:
                                    merged_cell_h1 = worksheet.cell(row=count_rows, column=4, value=e_header)
                                    merged_cell_h1.alignment = Alignment(horizontal='left')
                                else:
                                    if rows == 12:
                                        if '|' in e_header:
                                            pipe_index = e_header.index('|')  # Find the index of '|'
                                            e_value = e_header[:pipe_index]  # Extract 'E' before '|'
                                            pt_value = e_header[pipe_index + 1:]  # Extract 'PT' after '|'
                                            pt_value = pt_value.replace('|', '')
                                            combined_value = f"{e_value}{to_superscript(pt_value)}"
                                            # Set the combined value to the cell
                                            e_header = combined_value
                                    merged_cell_h1 = worksheet.cell(row=count_rows, column=start_header_columns,
                                                                    value=e_header)
                                    merged_cell_h1.alignment = Alignment(horizontal='center')
                                merged_cell_h1.font = small_font
                                merged_cell_h1.border = border
                            count_rows += 1
                gpa = sum_grade_point_credit / total_credit_hrs_taken
                gpa = math.floor(gpa * 10) / 10

                status = 'Continuing'
                if has_incomplete_course:
                    status = 'Incomplete'
                elif total_failed_core_subject > 0:
                    status = 'Probation'
                merged_cell_h1 = worksheet.cell(row=count_rows, column=1, value="GPA")
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border
                merged_cell_h1 = worksheet.cell(row=count_rows, column=2, value=gpa)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border

                worksheet.merge_cells(start_row=count_rows, start_column=3, end_row=count_rows, end_column=4)
                merged_cell_h1 = worksheet.cell(row=count_rows, column=3, value="Cumulative GPA")
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border
                merged_cell_h1 = worksheet.cell(row=count_rows, column=5, value=gpa)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border

                worksheet.merge_cells(start_row=count_rows, start_column=6, end_row=count_rows, end_column=7)
                merged_cell_h1 = worksheet.cell(row=count_rows, column=6, value="Credits Acquired")
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border
                merged_cell_h1 = worksheet.cell(row=count_rows, column=8, value=total_credit_hrs_taken)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border

                worksheet.merge_cells(start_row=count_rows, start_column=9, end_row=count_rows, end_column=11)
                merged_cell_h1 = worksheet.cell(row=count_rows, column=9, value="Cumulative Credits Acquired")
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border
                merged_cell_h1 = worksheet.cell(row=count_rows, column=12, value=total_credit_hrs_taken)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border

                count_rows += 1
                merged_cell_h1 = worksheet.cell(row=count_rows, column=1, value="Status")
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border
                merged_cell_h1 = worksheet.cell(row=count_rows, column=2, value=status)
                merged_cell_h1.alignment = Alignment(horizontal='center', wrapText=True)
                merged_cell_h1.font = small_font_border

                count_rows += 3
                worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=6)
                cell = worksheet.cell(row=count_rows, column=1, value="KEY TO EXAMINATION GRADING ")
                cell.alignment = Alignment(horizontal='center', wrapText=True)
                cell.font = small_font_border
                count_rows += 1
                worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=6)
                cell = worksheet.cell(row=count_rows, column=1, value="GRADING OF EXAMINATION SCORE")
                cell.alignment = Alignment(horizontal='center', wrapText=True)
                cell.font = small_font_border
                cell.border = border

                worksheet.merge_cells(start_row=count_rows, start_column=8, end_row=count_rows, end_column=12)
                cell = worksheet.cell(row=count_rows, column=1,
                                      value="1 CREDIT HOUR=30 LECTURE HOURS OR 60 PRACTICAL HOUR")
                cell.alignment = Alignment(horizontal='center', wrapText=True)
                cell.font = small_font_border
                cell.border = border

                count_rows += 1
                signature_row = count_rows

                final_list = ['GRADE', 'SCORES', 'GRADE POINTS', 'CLASSIFICATION']
                cl = 0
                for h in final_list:
                    cl += 1
                    if cl == 4:
                        worksheet.merge_cells(start_row=count_rows, start_column=cl, end_row=count_rows, end_column=6)
                    cell = worksheet.cell(row=count_rows, column=cl, value=h)
                    cell.alignment = Alignment(horizontal='center', wrapText=True)
                    cell.font = small_font
                    cell.border = border

                grade_lists = [
                    ['A', '70-100', '4.5-5', 'Excellent'],
                    ['B+', '65-69', '4-4.4', 'Very Good'],
                    ['B', '60-64', '3.0-3.9', 'Good'],
                    ['C', '50-59', '2.0-2.9', 'Pass'],
                    ['D', '40-49', '1.0-1.9', 'Marginal Fail'],
                    ['E', '0-39', '0.0-0.9', 'Absolute Fail']
                ]
                # cl = 0
                for rw in grade_lists:
                    cl = 0
                    for h in rw:
                        cl += 1
                        if cl == 4:
                            worksheet.merge_cells(start_row=count_rows, start_column=cl, end_row=count_rows,
                                                  end_column=6)
                        cell = worksheet.cell(row=count_rows, column=cl, value=h)
                        cell.alignment = Alignment(horizontal='center', wrapText=True)
                        cell.font = small_font
                        cell.border = border
                    count_rows += 1

                column_no = 12
                # Decode the Base64 image
                base64_image = signature_data['data']['base64doc']
                decoded_image = base64.b64decode(base64_image)

                # Load the image into BytesIO
                image_stream = BytesIO(decoded_image)

                # Create an Image object from BytesIO
                img = Image(image_stream)

                # Calculate the column letter for the anchor cell (e.g., "B" for column number 2)
                anchor_column = openpyxl.utils.get_column_letter(10)

                # Set the row number for the anchor cell
                anchor_row = signature_row + 1

                # # Set the anchor cell for the image to the bottom center cell
                anchor_cell = f'{anchor_column}{anchor_row}'
                # print(img)
                # Add image to the worksheet and set the anchor cell
                worksheet.add_image(img, anchor_cell)
                # worksheet[anchor_cell].alignment = Alignment(horizontal='center', vertical='bottom')

                name_row = signature_row + 4
                colum_no = 12
                worksheet.merge_cells(start_row=name_row, start_column=8, end_row=name_row, end_column=12)
                summary_text = worksheet.cell(row=name_row, column=8, value=info.context.user.full_name)
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border

                # print(info.context.user.full_name)

                name_row += 1
                worksheet.merge_cells(start_row=name_row, start_column=8, end_row=name_row,
                                      end_column=colum_no)
                summary_text = worksheet.cell(row=name_row, column=8, value=headship['data'])
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border

            # worksheet.column_dimensions['A'].width = 10
            # worksheet.column_dimensions['I'].width = 10
            worksheet.column_dimensions['L'].width = 13

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.protection = Protection(locked=True)

                # Protect the worksheet to make cells not editable
                worksheet.protection.sheet = True

            worksheet.print_area = ''

            # Save the workbook to a BytesIO buffer
            file_buffer = BytesIO()
            workbook.save(file_buffer)
            file_buffer.seek(0)

            # Convert the Excel file to Base64 string
            file_data = file_buffer.getvalue()
            base64_data = base64.b64encode(file_data).decode()

            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Semester Exam Results Retrieved Successfully",
                data=ExcelFile(base64_data=base64_data, file_name=file_name.upper())
            )

            return Response(
                status=False,
                code=ResponseCode.FAILURE,
                message="Please upload your signature before downloading Semester results",
                data=ExcelFile(base64_data=[], file_name=""),
            )

    @staticmethod
    def get_student_exam_results(student_uid) -> List[ExamResult]:
        with session_scope() as session:
            result = session.query(ExamResult).filter(ExamResult.student_uid == student_uid,
                                                      ExamResult.deleted_at.is_(None)).all()
            return result

    @staticmethod
    def publish_exam_result(program_semester_uids, info) -> Response[None]:
        with session_scope() as session:
            user_unit_department_uids = get_user_unit_department_headship(info)
            program_courses = session.query(ProgramCourse.id, ProgramCourse.forward_status).join(ProgramSemester).join(
                Program).filter(
                ProgramSemester.uid.in_(program_semester_uids),
                ProgramCourse.forward_status == 3,
                ProgramCourse.deleted_at.is_(None),
                ProgramSemester.deleted_at.is_(None),
                Program.department_uid.in_(user_unit_department_uids),
                Program.deleted_at.is_(None)).all()
            if not program_courses:
                return Response(
                    status=True,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="No Any Examination Is ready for Publishing from selected Programs",
                    data=None
                )

            forward_logs = []
            total = 0
            staff_uid = str(info.context.user.staff.uid)
            for pc in program_courses:
                status = pc.forward_status + 1
                total += 1
                logs = ExamCourseResultForwardLogs(
                    program_course_id=pc.id,
                    staff_uid=staff_uid,
                    staff_name=info.context.user.full_name,
                    forwarded_from=pc.forward_status,
                    forwarded_to=status
                )
                forward_logs.append(logs)
                session.query(ProgramCourse).filter_by(id=pc.id).update({"forward_status": status})
            session.add_all(forward_logs)
            session.commit()
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Program Examination Published Successful",
                data=None
            )

    @staticmethod
    def publish_all_exam_results(academic_year_uid,semester, info) -> Response[None]:
        with session_scope() as session:
            academic_year = AcademicYearService(AcademicYear).get(academic_year_uid)
            if academic_year is None:
                return Response(
                    status=False,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="Invalid Academic Year Selection",
                    data=None
                )
            # user_unit_department_uids = get_user_unit_department_headship(info)
            program_courses = session.query(ProgramCourse.id, ProgramCourse.forward_status).join(ProgramSemester).join(
                Program).filter(
                # ProgramSemester.uid.in_(program_semester_uids),
                ProgramSemester.academic_year_id == academic_year.id,
                ProgramSemester.semester == semester,
                ProgramCourse.forward_status == 3,
                ProgramCourse.deleted_at.is_(None),
                ProgramSemester.deleted_at.is_(None),
                # Program.department_uid.in_(user_unit_department_uids),
                Program.deleted_at.is_(None)).all()
            if not program_courses:
                return Response(
                    status=True,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="No Any Examination Is ready for Publishing from selected Programs",
                    data=None
                )

            forward_logs = []
            total = 0
            staff_uid = str(info.context.user.staff.uid)
            for pc in program_courses:
                status = pc.forward_status + 1
                total += 1
                logs = ExamCourseResultForwardLogs(
                    program_course_id=pc.id,
                    staff_uid=staff_uid,
                    staff_name=info.context.user.full_name,
                    forwarded_from=pc.forward_status,
                    forwarded_to=status
                )
                forward_logs.append(logs)
                session.query(ProgramCourse).filter_by(id=pc.id).update({"forward_status": status})
            session.add_all(forward_logs)
            session.commit()
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Program Examination Published Successful",
                data=None
            )

    @staticmethod
    def un_publish_all_semester_exam_result(academic_year_uid, semester, info) -> Response[None]:
        with session_scope() as session:
            academic_year = AcademicYearService(AcademicYear).get(academic_year_uid)
            if academic_year is None:
                return Response(
                    status=False,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="Invalid Academic Year Selection",
                    data=None
                )
            program_courses = session.query(ProgramCourse).join(ProgramSemester)\
                .filter(ProgramSemester.academic_year_id == academic_year.id, ProgramCourse.forward_status == 4,
                        ProgramSemester.semester == semester,ProgramCourse.deleted_at.is_(None), ProgramSemester.deleted_at.is_(None)).all()

            if not program_courses:
                return Response(
                    status=True,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message=f"No Any Examination Is Published in {academic_year.name} - Semester {semester}",
                    data=None
                )
            forward_logs = []
            total = 0
            staff_uid = str(info.context.user.staff.uid)
            for pc in program_courses:
                status = pc.forward_status - 1
                total += 1
                logs = ExamCourseResultForwardLogs(
                    program_course_id=pc.id,
                    staff_uid=staff_uid,
                    staff_name=info.context.user.full_name,
                    forwarded_from=pc.forward_status,
                    forwarded_to=status
                )
                forward_logs.append(logs)
                session.query(ProgramCourse).filter_by(id=pc.id).update({"forward_status": status})
            session.add_all(forward_logs)
            session.commit()
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Program Examination Unpublished Successful",
                data=None
            )

    @staticmethod
    def un_publish_semester_exam_result_by_program_semester_uids(program_semester_uids, info) -> Response[None]:
        with session_scope() as session:
            program_courses = session.query(ProgramCourse.id, ProgramCourse.forward_status).join(ProgramSemester).join(
                Program).filter(
                ProgramSemester.uid.in_(program_semester_uids),
                ProgramCourse.forward_status == 4,
                ProgramCourse.deleted_at.is_(None),
                ProgramSemester.deleted_at.is_(None),
                Program.deleted_at.is_(None)).all()
            if not program_courses:
                return Response(
                    status=True,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message=f"No Any Examination Is Published For the selected Program Semester",
                    data=None
                )
            forward_logs = []
            total = 0
            staff_uid = str(info.context.user.staff.uid)
            for pc in program_courses:
                status = pc.forward_status - 1
                total += 1
                logs = ExamCourseResultForwardLogs(
                    program_course_id=pc.id,
                    staff_uid=staff_uid,
                    staff_name=info.context.user.full_name,
                    forwarded_from=pc.forward_status,
                    forwarded_to=status
                )
                forward_logs.append(logs)
                session.query(ProgramCourse).filter_by(id=pc.id).update({"forward_status": status})
            session.add_all(forward_logs)
            session.commit()
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Program Examination Un Published Successful",
                data=None
            )
