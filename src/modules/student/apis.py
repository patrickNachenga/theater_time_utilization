import base64
import io
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, Protection
from io import BytesIO

import strawberry
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.core.security import CustomPermissionExtension, LoginRequiredExtension
from src.db.session import session_scope
from src.helpers.utils import get_current_academic_year, get_student_from_uaa, get_student_from_uaa_by_reg_numbers, insert_exam_result, insert_course_work, \
    general_upload
from src.models import ExamCategory, ProgramCourse
from src.modules.student.service import StudentService
from src.modules.exam_category.service import ExamCategoryService
from src.shared.response import Response
from src.shared.response_code import ResponseCode
from src.types import CourseRegistrationListNode, \
    CourseRegistrationInputNode, UaaDataResponse, StudentUaaData, ExcelFile, ProgramCourseListNode, \
    CourseRegisterInputNode, StudentProgramCourseListNode, ExamRegistrationInput, ExamRegistrationListNode, \
    ExamToRegister, ExamRegistrationNode, FailedStudent, UploadInput, UploadResponse


@strawberry.type
class StudentQuery:
    @strawberry.field(extensions=[LoginRequiredExtension()])
    def get_student_course_to_register(self, inputs: CourseRegisterInputNode) -> Response[StudentProgramCourseListNode]:
        try:
            result = StudentService().get_student_course_to_register(inputs)
            # print(result)
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Program courses Retrieved successfully",
                data=result)
        except Exception as e:
            print(e)
            # result = [CourseRegistrationListNode(course_to_register=None, total_count=0, course_registered=None)]
            return Response(
                status=False,
                code=ResponseCode.NO_RECORD_FOUND,
                message="Program courses not found",
                data=[])

    @strawberry.field(extensions=[LoginRequiredExtension()])
    def get_student_current_course_registration(self, student_uid: str) -> Response[CourseRegistrationListNode]:
        try:
            result = StudentService().get_student_current_course_registration(student_uid)

            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Course Registration Retrieved successfully",
                data=result)
        except Exception as e:
            print(e)
            result = CourseRegistrationListNode(items=[], total_count=0)
            return Response(
                status=False,
                code=ResponseCode.NO_RECORD_FOUND,
                message="Course Registration not found",
                data=result)

    @strawberry.field()
    def get_allocation_students(self, allocation_uid: str, assessment_number: int, exam_category_id: int,
                                out_off: int) -> UaaDataResponse:
        try:
            result = StudentService().get_allocation_students(allocation_uid, assessment_number, exam_category_id,
                                                              out_off,'REGNO')

            if result:
                response = UaaDataResponse(
                    status=True,
                    code=ResponseCode.SUCCESS,
                    message="Successfully Retrieved",
                    data=[StudentUaaData(registration_number=item['registration_number'], full_name=StudentQuery.reformat_name(item.get('user', {}).get('full_name')),
                                         uid=item['uid'], score=item['marks'])
                          for item in result['data']]
                )

                return response
            else:
                return UaaDataResponse(
                    status=False,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="Failed to retrieve",
                    data=[])
        except Exception as e:
            print(e)
            return UaaDataResponse(
                status=False,
                code=ResponseCode.NO_RECORD_FOUND,
                message="Failed to retrieve",
                data=[])

    @staticmethod
    def reformat_name(full_name: str) -> str:
        # Split the input string into parts
        parts = full_name.split()

        # Capitalize the first part
        parts[0] = parts[0].upper() + ', '

        # Capitalize the first letter of the second part
        parts[1] = parts[1][0].upper() + parts[1][1:].lower()

        # Capitalize the first letter of the third part
        if len(parts) > 2:
            parts[2] = parts[2][0].upper() + parts[2][1:].lower()

        # Join the parts back into a single string
        full_name_ = ' '.join(parts)

        return full_name_
    @strawberry.field
    def get_student_current_registered_exam(self, student_uid: str) -> Response[List[ExamRegistrationNode]]:
        try:
            result = StudentService().get_student_current_registered_exam(student_uid)
            if result:
                return Response(
                    status=True,
                    code=ResponseCode.SUCCESS,
                    message="Exam Registration Retrieved successfully",
                    data=result)
            else:
                return Response(
                    status=False,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="Exam Registration not found",
                    data=[])
        except Exception as e:
            print(e)
            return Response(
                status=False,
                code=ResponseCode.NO_RECORD_FOUND,
                message="Exam Registration not found exception occurred",
                data=[])

    @strawberry.field
    def get_student_exam_to_register(self, student_uid: str) -> Response[ExamToRegister]:
        try:
            result = StudentService().get_student_exam_to_register(student_uid)
            if result:
                return Response(
                    status=True,
                    code=ResponseCode.SUCCESS,
                    message="Exam Registration Retrieved successfully",
                    data=result)
            else:
                return Response(
                    status=False,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="Exams not found",
                    data=[])
        except Exception as e:
            print(e)
            return Response(
                status=False,
                code=ResponseCode.NO_RECORD_FOUND,
                message="Exams not found",
                data=[])


@strawberry.type
class StudentMutation:
    @strawberry.field(extensions=[CustomPermissionExtension(["REGISTER_STUDENT_COURSES"])])
    def register_student_course(self, inputs: List[CourseRegistrationInputNode], remove: List[str]) -> Response[
        CourseRegistrationListNode]:
        try:
            result = StudentService().register_student_course(inputs, remove)
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Course Registered successfully",
                data=result)
        except Exception as e:
            print(e)
            result = CourseRegistrationListNode(items=[], total_count=0)
        return Response(status=False, code=ResponseCode.FAILURE, message="Failed to register course", data=result)

    @strawberry.field
    def generate_allocation_xls_template(self, allocation_uid: str, out_off: int, exam_category_id: int,
                                         assessment_number: int, assessment_weight: int, excel_sorting:str = 'REGNO') -> ExcelFile:
        result = StudentService().get_allocation_students(allocation_uid, assessment_number, exam_category_id, out_off, excel_sorting)
        file_buffer = io.BytesIO()

        file_name = ''
        # Create a new workbook
        workbook = Workbook()

        # Create a new worksheet
        worksheet = workbook.active
        # Set column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 45

        # Set the font style to Times New Roman
        font = Font(name="Times New Roman")
        font_border = Font(name="Times New Roman", bold=True)
        # Set the border style
        border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))

        # Define the vertical headers
        vertical_headers = ["Course Ante", "Program Code", "Academic Year", "Study Year", "Exam Category",
                            "Assessment No", "Mark Out of",
                            "Assessment Weight"]

        file_name += result["program_course"].program_semester.program.code + "_"
        file_name += result["program_course"].course.code + "_"

        # Sample data for the vertical header
        data = {
            "Course Ante": result["program_course"].course.code,
            "Program Code": result["program_course"].program_semester.program.code,
            "Academic Year": result["program_course"].program_semester.academic_year.name,
            "Study Year": str(result["program_course"].program_semester.study_year),
            "Exam Category": str(exam_category_id),
            "Assessment No": str(assessment_number),
            "Mark Out of": str(out_off),
            "Assessment Weight": str(assessment_weight)
        }
        worksheet.sheet_view.showGridLines = False
        # Generate the data for the vertical header
        vertical_data = [data[header] for header in vertical_headers]
        for row, header in enumerate(vertical_headers, start=1):
            cell = worksheet.cell(row=row, column=1, value=header)
            cell.alignment = Alignment(horizontal='left')
            cell.font = font_border
            cell.border = None
            cell.protection = Protection(locked=False)
        for row, value in enumerate(vertical_data, start=1):
            cell = worksheet[f"C{row}"]
            cell.value = value
            cell.font = font_border
            cell.border = None
            cell.protection = Protection(locked=False)
        # Define the horizontal headers
        # worksheet.sheet_view.showGridLines = True
        horizontal_headers = ["SN", "Reg No", "Name", "Marks"]
        start_col = 1  # Start column for horizontal headers
        start_row = len(vertical_headers) + 1  # Start row for horizontal headers

        for col, header in enumerate(horizontal_headers, start=start_col):
            cell = worksheet.cell(row=start_row, column=col, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font_border
            cell.border = border

        count = 0

        for row, item in enumerate(result['data'], start=10):

            full_name = item.get('user', {}).get('full_name')

            # Split the input string into parts
            parts = full_name.split()

            # Capitalize the first part
            parts[0] = parts[0].upper() + ', '

            # Capitalize the first letter of the second part
            parts[1] = parts[1][0].upper() + parts[1][1:].lower()

            # Capitalize the first letter of the third part
            if len(parts) > 2:
                parts[2] = parts[2][0].upper() + parts[2][1:].lower()

            # Join the parts back into a single string
            full_name_ = ' '.join(parts)





            count += 1
            worksheet[f"A{row}"] = count
            worksheet[f"B{row}"] = item['registration_number']
            # worksheet[f"C{row}"] = item.get('user', {}).get('full_name')
            worksheet[f"C{row}"] = full_name_
            # worksheet[f"C{row}"] = item['last_name'].capitalize() + ", " + item['first_name'] + " " + item['middle_name']
            worksheet[f"D{row}"] = item['marks']
            if isinstance(item['marks'], (int, float)):
                worksheet[f"D{row}"] = round(item['marks'] * 10) / 10
            # Align the cells to the center
            for col in range(1, 5):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = font
                cell.border = border
        worksheet.cell(row=1, column=4, value=str(result["program_course"].id))
        worksheet.cell(row=1, column=4).font = font_border


        worksheet.cell(row=1, column=5, value=str(result["program_course"].course.name))
        worksheet.cell(row=1, column=5).font = font_border

        # get exam category

        exam_cat_result = ExamCategoryService().get_exam_categories_by_id(exam_category_id)
        for row in exam_cat_result:
            file_name += str(row.code + "_" + str(assessment_number))
            exam_cat_name = row.name
            worksheet.cell(row=2, column=5, value=str(exam_cat_name + ": " + str(assessment_number)))
            worksheet.cell(row=2, column=5).font = font_border

        # Set the specific column where cells should be non-editable (except column D)
        editable_column = 'D'

        # Iterate over rows in the worksheet
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                # Check if the current column is the editable column and cell row is not greater than 9
                if cell.column_letter == editable_column and cell.row >= 10:
                    # Set protection to False for the editable column
                    cell.protection = Protection(locked=False)
                    cell.number_format = '0.00'
                else:
                    # Set protection to True for other columns
                    cell.protection = Protection(locked=True)

            # Protect the worksheet to make cells not editable
            worksheet.protection.sheet = True

        # Save the workbook
        # workbook.save("layout.xlsx")

        # Save the workbook to a BytesIO buffer
        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        # Convert the Excel file to Base64 string
        file_data = file_buffer.getvalue()
        base64_data = base64.b64encode(file_data).decode()


        # Return the Base64 string as the result
        return ExcelFile(base64_data=base64_data, file_name=file_name)

    def reformat_name(full_name: str):
        # Split the input string into parts
        parts = full_name.split()

        # Capitalize the first part
        parts[0] = parts[0].upper()

        # Capitalize the first letter of the second part
        parts[1] = parts[1][0].upper() + parts[1][1:]

        # Capitalize the first letter of the third part

        if len(parts) > 2:
            parts[2] = parts[2][0].upper()

        # Join the parts back into a single string
        modified_string = ' '.join(parts)

        return modified_string

    @strawberry.field
    def register_student_exam(self, inputs: List[ExamRegistrationInput]) -> Response[ExamRegistrationListNode]:
        try:
            result = StudentService().register_student_exam(inputs)
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Exam Registered successfully",
                data=result)
        except Exception as e:
            print(e)
            result = ExamRegistrationListNode(items=[], total_count=0)
        return Response(status=False, code=ResponseCode.FAILURE, message="Failed to register exam", data=result)

    @strawberry.mutation
    async def upload_score(self, base64_file: str) -> Response[UploadResponse]:

        # Decode the base64 file content
        print("1")
        file_content = base64.b64decode(base64_file)
        print("2")
        # Load the workbook from the file content
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        # Get the desired worksheet by name or index
        worksheet = workbook.active  # Modify this line with the appropriate worksheet name or index
        exam_category_id = worksheet.cell(row=5, column=3).value
        assessment_number = worksheet.cell(row=6, column=3).value
        out_off = float(worksheet.cell(row=7, column=3).value)
        program_course_id = worksheet.cell(row=1, column=4).value
        print("3")
        # if program_course_id != 407:
        #     return Response(status=False, code=ResponseCode.FAILURE,
        #                     message="Please try after 1 hour (1400), there is upgrade ongoing", data=UploadResponse(
        #             success=0,
        #             failed=0,
        #             failed_students=[],
        #             success_students=[]
        #         ))

        weight = float(worksheet.cell(row=8, column=3).value)

        # Assuming the data is in a specific sheet and columns

        reg_no_column = 2  # Assuming Reg No is in column B

        marks_column = 4  # Assuming Marks is in column D

        with session_scope() as session:
            print("4")
            is_ue = session.query(ExamCategory.is_ue).filter(
                ExamCategory.id == exam_category_id).first().is_ue
            # get student list from uaa service to get student uid after filtering

            # get list of registration numbers from the template
            reg_numbers = []  # Initialize an empty list to store registration numbers
            for row in worksheet.iter_rows(min_row=10, values_only=True):
                reg_number_ = row[reg_no_column - 1]
                reg_numbers.append(reg_number_)  # Append the registration number to the list

            # Now reg_numbers contains all the registration numbers from the specified worksheet rows
            print("5")
            students = get_student_from_uaa_by_reg_numbers(reg_numbers)

            print("6")
            success = 0
            failed = 0
            failed_students = []
            success_students = []

            for row in worksheet.iter_rows(min_row=10, values_only=True):
                reg_number = row[reg_no_column - 1]
                # Find the item with the specified registration_number
                if row[marks_column - 1] is None:
                    continue
                try:
                    score = float(row[marks_column - 1])
                except ValueError:
                    score = 'InvalidMarks'
                    # print("Could not convert string to float." )


                program_course = session.query(ProgramCourse).filter(ProgramCourse.id == program_course_id,
                                                                     ProgramCourse.deleted_at.is_(None)).first()

                exam_category = session.query(ExamCategory).filter(ExamCategory.id == exam_category_id,
                                                                       ExamCategory.deleted_at.is_(None)).first()

                success_, failed_, failed_student, success_student = general_upload(session = session, students=students,
                                                                   program_course_id=program_course_id,
                                                                   exam_category_id=exam_category_id, score=score,
                                                                   out_off=out_off, weight=weight, is_ue=is_ue,
                                                                   reg_number=reg_number,
                                                                   assessment_number=assessment_number,
                                                                    program_course=program_course,
                                                                    exam_category = exam_category
                                                                    )
                success = success + success_
                failed = failed + failed_
                if failed_student.reg_number is not None:
                    failed_students.append(failed_student)
                if success_student.reg_number is not None:
                    success_students.append(success_student)

            # session.commit()
            response_data = UploadResponse(
                success=success,
                failed=failed,
                failed_students=failed_students,
                success_students=success_students
            )

            return Response(status=True, code=ResponseCode.SUCCESS, message="Executed successfully", data=response_data)

    @strawberry.mutation
    async def upload_online_score(self, inputs: UploadInput) -> Response[UploadResponse]:
        try:
            result = StudentService().upload_online_score(inputs)
            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Executed successfully",
                data=result)
        except Exception as e:
            print(e)
            result = []
        return Response(status=False, code=ResponseCode.FAILURE, message="Failed to execute", data=result)
