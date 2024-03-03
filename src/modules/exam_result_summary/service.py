import base64
import json
from io import BytesIO
from typing import List, Type
import requests

import openpyxl
from fastapi import APIRouter, UploadFile, File
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, Protection, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from src.api_routes.program_api import reformat_name
from src.core.config import settings
from src.core.security import Info
from src.db.session import session_scope
from src.helpers.utils import can_progress
from src.models import ExamResultSummary, Process, Workflow, State, ProcessFlow, StudentCourseRegistration
from src.modules import CRUDBase
from src.modules.academic_year.service import AcademicYearCrud
from src.modules.program_course.service import ProgramCourseCrud
from src.modules.program_semester.service import ProgramSemesterCrud
from src.modules.programs.service import ProgramService
from src.shared.response import Response
from src.shared.response_code import ResponseCode
from src.types import ExamResultSummaryInput, ExamResultSummarySearchCriteria, ExcelFile
from openpyxl import Workbook
import math


class ExamResultSummaryService((CRUDBase[ExamResultSummary, ExamResultSummaryInput, ExamResultSummaryInput])):
    @staticmethod
    def get_exam_result_summaries(search_criteria: ExamResultSummarySearchCriteria) -> List[ExamResultSummary]:
        with session_scope() as session:
            query = session.query(ExamResultSummary).filter(ExamResultSummary.deleted_at.is_(None))

            if search_criteria.gender:
                query = query.filter(ExamResultSummary.gender == search_criteria.gender)
            if search_criteria.program_course_id:
                query = query.filter(ExamResultSummary.program_course_id == search_criteria.program_course_id)
            if search_criteria.student_uid:
                query = query.filter(ExamResultSummary.student_uid == search_criteria.student_uid)
            if search_criteria.registration_number:
                query = query.filter(ExamResultSummary.registration_number == search_criteria.registration_number)
            if search_criteria.course_code:
                query = query.filter(ExamResultSummary.course_code == search_criteria.course_code)
            if search_criteria.academic_year_uid:
                query = query.filter(ExamResultSummary.academic_year_uid == search_criteria.academic_year_uid)
            if search_criteria.program_uid:
                query = query.filter(ExamResultSummary.program_uid == search_criteria.program_uid)
            if search_criteria.course_category:
                query = query.filter(ExamResultSummary.course_category == search_criteria.course_category)
            if search_criteria.semester:
                query = query.filter(ExamResultSummary.semester == search_criteria.semester)
            query = query.order_by(ExamResultSummary.registration_number.asc())
            # Execute the final query
            results = query.all()
            return results

    @staticmethod
    def generate_semester_exam_results(program_uid: str, academic_year_uid: str, semester: int,
                                       year_of_study: int, info: Info) -> Response[ExcelFile]:
        with (session_scope() as session):
            # result = StudentService().get_allocation_students(allocation_uid)
            # Create a new workbook
            workbook = Workbook()
            program = ProgramService.get_program_by_uid(program_uid)
            if program is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Program is not found",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            academic_year = AcademicYearCrud.get_academic_year_by_uid(academic_year_uid)
            if academic_year is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Academic Year Is Not Found",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            # Get Program Semester Information
            program_semester = ProgramSemesterCrud.get_program_semester_by_data(semester=semester,
                                                                                year_of_study=year_of_study,
                                                                                program_id=program.id,
                                                                                academic_year_id=academic_year.id)
            if not program_semester:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Program Semester Is Not Defined",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            if info.context.user is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Session Expired",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            headers = {
                "Content-Type": "application/json"
            }
            data_obj = {
                "user_uid": str(info.context.user.uid),
                "document_type": "signature"
            }
            payload = json.dumps(data_obj)
            # Send the POST request
            response = requests.post(settings.UAA_URi + f'/user/document?user_uid={info.context.user.uid}&document_type=signature', headers=headers)
            response.raise_for_status()
            if response.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Get User Signature",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            signature_data = response.json()
            if not signature_data['status']:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message=signature_data['message'],
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            get_headship = requests.get(
                settings.UAA_URi + f'/user/check-user-headship?user_uid={info.context.user.uid}&'
                                   f'department_uid={program.department_uid}', headers=headers)
            get_headship.raise_for_status()
            if get_headship.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Find Headship",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            headship = get_headship.json()
            if headship['status'] != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message=headship['message'],
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            # print(signature_data)
            file_name = f"{program.name}({program.code}) {academic_year.name} - YEAR {year_of_study} SEMESTER {semester} EXAMINATION RESULT"
            # Create a new worksheet
            worksheet = workbook.active
            # Set column widths
            # worksheet.column_dimensions['A'].width = 15
            # worksheet.column_dimensions['B'].width = 20
            # worksheet.column_dimensions['C'].width = 45

            # Set the font style to Times New Roman
            font = Font(name="Times New Roman", size=12)
            small_font = Font(name="Times New Roman", size=10)
            fill_color = PatternFill(start_color='FF999999', end_color='FF999999', fill_type='solid')
            font_border = Font(name="Times New Roman", bold=True, size=12)
            small_font_border = Font(name="Times New Roman", bold=True, size=10)
            # Set the border style
            border = Border(left=Side(border_style="thin"),
                            right=Side(border_style="thin"),
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))

            # Merge cells from C1 to J1
            worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
            # Merge cells from C2 to J2
            worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=10)

            merged_cell_h1 = worksheet.cell(row=1, column=3, value="SOKOINE UNIVERSITY OF AGRICULTURE")
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = font
            # Merge cells from C2 to J2
            merged_cell_h2 = worksheet.cell(row=2, column=3, value="P.O. BOX 3000, CHUO KIKUU, MOROGORO, TANZANIA")
            merged_cell_h2.alignment = Alignment(horizontal='center')
            merged_cell_h2.font = font
            count_rows = 4

            # Define the vertical headers
            vertical_headers = ["Semester Result", "Program", "Program Code", "Academic Year", "Study Year", "Semester",
                                "Semester Credit"]
            # Sample data for the vertical header
            data = {
                "Semester Result": "",
                "Program": str(program.name),
                "Program Code": str(program.code),
                "Academic Year": str(academic_year.name),
                "Study Year": str(year_of_study),
                "Semester": str(semester),
                "Semester Credit": str(program_semester.core_credits + program_semester.elective_credits)
            }
            worksheet.sheet_view.showGridLines = False

            # Generate the data for the vertical header
            vertical_data = [data[header] for header in vertical_headers]
            for row, header in enumerate(vertical_headers, start=4):
                count_rows += 1
                cell = worksheet.cell(row=row, column=3)
                cell.value = header
                cell.alignment = Alignment(horizontal='left')
                cell.font = font_border
                cell.border = None
                cell.protection = Protection(locked=False)
            for row, value in enumerate(vertical_data, start=4):
                cell = worksheet[f"F{row}"]
                cell.value = value
                cell.font = font_border
                cell.border = None
                cell.protection = Protection(locked=False)

            # Define the horizontal headers
            # worksheet.sheet_view.showGridLines = True
            start_col = 1  # Start column for horizontal headers

            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=16)
            summary_text = worksheet.cell(row=count_rows, column=1, value="Summary")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border
            count_rows += 1
            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=2)
            summary_text = worksheet.cell(row=count_rows, column=1, value="Status")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            summary_text = worksheet.cell(row=count_rows, column=3, value="Male")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            summary_text = worksheet.cell(row=count_rows, column=4, value="Female")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=5, end_row=count_rows, end_column=7)
            summary_text = worksheet.cell(row=count_rows, column=5, value="Total")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=8, end_row=count_rows, end_column=10)
            summary_text = worksheet.cell(row=count_rows, column=8, value="Male %")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=11, end_row=count_rows, end_column=13)
            summary_text = worksheet.cell(row=count_rows, column=11, value="Female %")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=14, end_row=count_rows, end_column=16)
            summary_text = worksheet.cell(row=count_rows, column=14, value="Total %")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            summary_text.border = border
            count_rows += 1

            status_info = ["CONTINUING", "PROBATION", "INCOMPLETE", "POSTPONE", "RETAKE", "TOTAL"]
            status_data = []
            for status in status_info:
                s_info = {
                    status: {
                        'male': 0,
                        'female': 0,
                        'total': 0,
                        'male%': 0,
                        'female%': 0,
                        'total%': 0,
                    }
                }
                status_data.append(s_info)
                worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows, end_column=2)
                cell = worksheet.cell(row=count_rows, column=1, value=status)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = font_border
                cell.border = border
                count_rows += 1

            # skip status

            # return status
            # Get Courses
            total_elective_courses = 0
            total_core_courses = 0
            program_courses = ProgramCourseCrud.get_program_course_by_program_semester(
                program_semester_id=program_semester.id)
            if program_courses is not None:
                courses_list = ""
                for pc in program_courses:
                    if pc.course_category.name.upper() == 'CORE':
                        total_core_courses += 1
                    elif pc.course_category.name.upper() == 'ELECTIVE':
                        total_elective_courses += 1
                    courses_list += f"{pc.course.code}: {pc.course.name}, "

                worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows,
                                      end_column=len(program_courses) + 11)
                cell = worksheet.cell(row=count_rows, column=1, value=courses_list)
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True, indent=1)
                cell.font = font
                cell.border = border
                # Calculate the number of lines in the wrapped text
                text_length = len(courses_list)
                num_lines = max(1,
                                text_length // 100)  # Assuming an average line width of 100 characters, adjust as needed
                # Set the row height based on the number of lines
                row_height_per_line = 15  # Adjust as needed
                worksheet.row_dimensions[count_rows].height = num_lines * row_height_per_line
                count_rows += 1

                course_colum_merge = 5
                worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows,
                                      end_column=course_colum_merge)
                summary_text = worksheet.cell(row=count_rows, column=1, value="Course")
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border
                summary_text.border = border

                # Core Course Title
                worksheet.merge_cells(start_row=count_rows, start_column=course_colum_merge + 1, end_row=count_rows,
                                      end_column=course_colum_merge + total_core_courses)
                summary_text = worksheet.cell(row=count_rows, column=course_colum_merge + 1, value="Core")
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border
                summary_text.border = border
                course_colum_merge += total_core_courses
                # course_colum_merge += 1

                # print("S:", total_core_courses)
                # print("Sum of Elective Courses:", total_elective_courses)
                # Elective Course Title
                worksheet.merge_cells(start_row=count_rows, start_column=course_colum_merge + 1, end_row=count_rows,
                                      end_column=total_elective_courses + course_colum_merge)
                summary_text = worksheet.cell(row=count_rows, column=course_colum_merge + 1, value="Elective")
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border
                summary_text.border = border
                count_rows += 1

                credit_hr_colum_merge = 5
                worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows,
                                      end_column=credit_hr_colum_merge)
                summary_text = worksheet.cell(row=count_rows, column=1, value="Credit Hours")
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border
                summary_text.border = border

                for pc in program_courses:
                    credit_hr_colum_merge += 1
                    cel = worksheet.cell(row=count_rows, column=credit_hr_colum_merge, value=pc.credit)
                    cel.alignment = Alignment(horizontal='center')
                    cel.font = font_border
                    cel.border = border

                count_rows += 1

                # SN
                summary_text = worksheet.cell(row=count_rows, column=1, value="SNo")
                summary_text.alignment = Alignment(horizontal='center')
                summary_text.font = font_border
                summary_text.border = border
                # Name
                text = worksheet.cell(row=count_rows, column=2, value="Name")
                text.alignment = Alignment(horizontal='center')
                text.font = font_border
                text.border = border

                # SN
                text = worksheet.cell(row=count_rows, column=3, value="Reg. No.")
                text.alignment = Alignment(horizontal='center')
                text.font = font_border
                text.border = border

                text = worksheet.cell(row=count_rows, column=4, value="Sex")
                text.alignment = Alignment(horizontal='center')
                text.font = font_border
                text.border = border

                text = worksheet.cell(row=count_rows, column=5, value="Examination No.")
                text.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, indent=1)
                text.font = font_border
                text.border = border

                course_code_colum_merge = 5
                for pc in program_courses:
                    course_code_colum_merge += 1
                    cel = worksheet.cell(row=count_rows, column=course_code_colum_merge, value=pc.course.code)
                    cel.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
                    cel.font = font_border
                    cel.border = border
                    # Set the row height based on the number of lines
                    row_height_per_line = 10  # Adjust as needed
                    worksheet.row_dimensions[count_rows].height = 8 * row_height_per_line

                    column_letter = get_column_letter(course_code_colum_merge)
                    # Set the width of the column
                    worksheet.column_dimensions[column_letter].width = 3
                total_titles = ["TOTAL CREDITS TAKEN", "CREDITS ACQUIRED", "CORE COURSES FAILED", "GPA", "Remarks",
                                "Courses Under Probation"]
                for title in total_titles:
                    course_code_colum_merge += 1
                    start_column = course_code_colum_merge
                    end_column = course_code_colum_merge
                    column_letter = get_column_letter(start_column)

                    # Set the value of the cell before merging
                    cel = worksheet.cell(row=count_rows - 2, column=start_column, value=title)
                    if title == 'Remarks' or title == 'Courses Under Probation':
                        cel.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, indent=1)
                        column_letter = get_column_letter(start_column)
                        if title == 'Remarks':
                            worksheet.column_dimensions[column_letter].width = 10
                        else:
                            worksheet.column_dimensions[column_letter].width = 35

                    else:
                        cel.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
                        # Increment the course_code_colum_merge by the width of the merged area minus 1
                        course_code_colum_merge += (end_column - start_column)
                        # Set the row height based on the number of lines
                        row_height_per_line = 10  # Adjust as needed
                        worksheet.row_dimensions[count_rows].height = 12 * row_height_per_line
                        # Set the width of the column
                        worksheet.column_dimensions[column_letter].width = 6

                    cel.font = font_border
                    cel.border = border
                    # Merge cells after setting the value
                    worksheet.merge_cells(start_row=count_rows - 2, start_column=start_column, end_row=count_rows,
                                          end_column=end_column)
                # Get students list
                results = session.query(ExamResultSummary.student_uid, func.max(ExamResultSummary.gender).label('sex'),
                                        func.max(ExamResultSummary.registration_number).label('registration_number'),
                                        func.concat(func.max(ExamResultSummary.first_name), ' ',
                                                    func.max(ExamResultSummary.middle_name), ' ',
                                                    func.max(ExamResultSummary.last_name)).label('full_name')).filter(
                    ExamResultSummary.program_uid == program_uid,
                    ExamResultSummary.academic_year_uid == academic_year_uid,
                    ExamResultSummary.semester == semester, ExamResultSummary.study_year == year_of_study).group_by(
                    ExamResultSummary.student_uid).all()
                if len(results) > 0:
                    count_rows += 1
                    count = 0
                    for row, item in enumerate(results, start=count_rows):
                        count += 1
                        courses_under_probation = ""
                        total_credit_hrs_taken = 0
                        total_credit_hrs_acquired = 0
                        total_failed_core_subject = 0
                        sum_grade_point_credit = 0
                        failed_subjects = 0
                        passed_subjects = 0
                        remark_status = 0
                        # count_rows += 1
                        worksheet[f"A{row}"] = count
                        worksheet[f"B{row}"] = reformat_name(item['full_name'])
                        worksheet[f"C{row}"] = item['registration_number']
                        worksheet[f"D{row}"] = item['sex'][0]
                        worksheet[f"E{row}"] = program.registration_code
                        sex = item['sex'][0].upper()
                        for col in range(1, 6):
                            cell = worksheet.cell(row=row, column=col)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.font = small_font
                            cell.border = border
                        col = 5
                        for pc in program_courses:
                            col += 1
                            exam_result = session.query(ExamResultSummary.grade, ExamResultSummary.grade_remark, ExamResultSummary.grade_point_credit).filter(
                                ExamResultSummary.student_uid == item['student_uid'],
                                ExamResultSummary.program_course_id == pc.id).first()
                            if exam_result:
                                if exam_result.grade_point_credit is not None:
                                    sum_grade_point_credit += exam_result.grade_point_credit
                                value = exam_result.grade
                                # Check if student have passed this course sum
                                if exam_result.grade_remark.upper() == "PASS":
                                    total_credit_hrs_acquired += pc.credit
                                    passed_subjects += 1
                                else:
                                    if pc.course_category.name.upper() == 'CORE' and value != 'I':
                                        total_failed_core_subject += 1
                                    if value != 'I':
                                        courses_under_probation += pc.course.code + ", "
                                        failed_subjects += 1
                            else:
                                value = '-'
                            cel = worksheet.cell(row=count_rows, column=col, value=value)
                            cel.alignment = Alignment(horizontal='center')
                            cel.font = small_font
                            if exam_result:
                                if value != 'I' and exam_result.grade_remark.upper() != "PASS":
                                    cel.font = small_font_border
                                    cel.fill = fill_color
                            cel.border = border
                            # Check if student have registered this course sum its credit to total_credit_hrs_taken
                            if session.query(StudentCourseRegistration.id).filter(
                                    StudentCourseRegistration.student_uid == item['student_uid'],
                                    StudentCourseRegistration.program_course_id == pc.id).first():
                                total_credit_hrs_taken += pc.credit

                        continuing_data = status_data[0]['CONTINUING']
                        probation_data = status_data[1]['PROBATION']
                        incomplete_data = status_data[2]['INCOMPLETE']
                        postpone_data = status_data[3]['POSTPONE']
                        retake_data = status_data[4]['RETAKE']
                        if failed_subjects > 0:
                            remark_status = 'PROBATION'
                            if sex == 'M':
                                probation_data['male'] += 1
                            else:
                                probation_data['female'] += 1
                        elif len(program_courses) != passed_subjects:
                            remark_status = 'INCOMPLETE'
                            if sex == 'M':
                                incomplete_data['male'] += 1
                            else:
                                incomplete_data['female'] += 1
                        else:
                            remark_status = 'CONTINUING'
                            if sex == 'M':
                                continuing_data['male'] += 1
                            else:
                                continuing_data['female'] += 1
                        # calculate the GPA

                        # gpa = '-'
                        # if remark_status != 'INCOMPLETE':
                        gpa = sum_grade_point_credit / total_credit_hrs_taken
                        gpa =  math.floor(gpa * 10) / 10
                        total_title_results = [total_credit_hrs_taken, total_credit_hrs_acquired,
                                               total_failed_core_subject, gpa, remark_status, courses_under_probation]

                        for title_ in total_title_results:
                            col += 1
                            cel = worksheet.cell(row=count_rows, column=col, value=title_)
                            cel.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, indent=1)
                            cel.font = small_font
                            cel.border = border
                        count_rows += 1
                    worksheet.column_dimensions['A'].width = 4
                    worksheet.column_dimensions['B'].width = 25
                    worksheet.column_dimensions['C'].width = 13
                    worksheet.column_dimensions['D'].width = 13
                else:
                    return Response(
                        status=False,
                        code=ResponseCode.FAILURE,
                        message=f"No Result Found For {file_name}",
                        data=ExcelFile(base64_data=[], file_name=""),
                    )
            simplified_status_data = []

            total_data = status_data[5]['TOTAL']
            for status in status_data:
                key = list(status.keys())[0].lower()
                values = status[key.upper()]
                if key.upper() != 'TOTAL':
                    total = values['male'] + values['female']
                    total_data['total'] += total
                    total_data['male'] += values['male']
                    total_data['female'] += values['female']
                    male_perc = round((values['male'] / count) * 100, 2)
                    total_data['male%'] += male_perc
                    female_perc = round((values['female'] / count) * 100, 2)
                    total_data['female%'] += female_perc
                    total_perc = male_perc + female_perc
                    total_data['total%'] += total_perc
                    simplified_status_data.append(
                        [values['male'], values['female'], total, male_perc, female_perc, total_perc])
                else:
                    simplified_status_data.append(
                        [total_data['male'], total_data['female'], total_data['total'], round(total_data['male%'], 2),
                         round(total_data['female%'], 2), round(total_data['total%'])])

            # print(simplified_status_data)
            signature_row = count_rows
            count_rows = 13
            # print(simplified_status_data)
            for status in simplified_status_data:
                # print("===== =======")
                start_column = 2
                end_column = 2
                status_column = 0
                for value in status:
                    start_column += 1
                    end_column += 1
                    status_column += 1
                    if status_column > 2:
                        end_column += 2
                        start_column = end_column - 2
                        # print(f"{count_rows} ====> {start_column} - {end_column}")
                        worksheet.merge_cells(start_row=count_rows, start_column=start_column, end_row=count_rows,
                                              end_column=end_column)
                    cell = worksheet.cell(row=count_rows, column=start_column, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                    cell.font = font_border
                    cell.border = border
                count_rows += 1
            # Iterate over rows in the worksheet


            column_no = len(program_courses) + 11
            # Decode the Base64 image
            base64_image = signature_data['data']['base64doc']
            decoded_image = base64.b64decode(base64_image)

            # Load the image into BytesIO
            image_stream = BytesIO(decoded_image)

            # Create an Image object from BytesIO
            img = Image(image_stream)

            # Calculate the coordinates of the bottom center cell
            bottom_center_cell = openpyxl.utils.coordinate_to_tuple(
                f'{openpyxl.utils.get_column_letter(column_no)}{signature_row+1}')

            # Calculate the column letter for the anchor cell (e.g., "B" for column number 2)
            anchor_column = openpyxl.utils.get_column_letter(bottom_center_cell[1]-13)

            # Set the row number for the anchor cell
            anchor_row = bottom_center_cell[0]

            print(anchor_column)
            print(bottom_center_cell[0])

            # Set the anchor cell for the image to the bottom center cell
            anchor_cell = f'{anchor_column}{anchor_row}'
            # Add image to the worksheet and set the anchor cell
            worksheet.add_image(img, anchor_cell)
            worksheet[anchor_cell].alignment = Alignment(horizontal='center', vertical='bottom')




            name_row = signature_row + 4
            colum_no = len(program_courses) + 11
            worksheet.merge_cells(start_row=name_row, start_column=1, end_row=name_row, end_column=colum_no)
            summary_text = worksheet.cell(row=name_row, column=1, value=info.context.user.full_name)
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border
            # print(info.context.user.full_name)


            name_row += 1
            worksheet.merge_cells(start_row=name_row, start_column=1, end_row=name_row,
                                  end_column=colum_no)
            summary_text = worksheet.cell(row=name_row, column=1, value=headship['data'])
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border




            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.protection = Protection(locked=True)

                # Protect the worksheet to make cells not editable
                worksheet.protection.sheet = True

            # Save the workbook
            # workbook.save("layout.xlsx")

            # Save the workbook to a BytesIO buffer
            file_buffer = BytesIO()
            workbook.save(file_buffer)
            file_buffer.seek(0)

            # Convert the Excel file to Base64 string
            file_data = file_buffer.getvalue()
            base64_data = base64.b64encode(file_data).decode()

            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Semester Exam Results Retrieved Successfully",
                data=ExcelFile(base64_data=base64_data, file_name=file_name.upper())
            )

    @staticmethod
    def get_student_exam_result_summaries(student_uid: str) -> List[ExamResultSummary]:
        with session_scope() as session:
            result = session.query(ExamResultSummary).filter(ExamResultSummary.student_uid == student_uid,
                                                             ExamResultSummary.publish_status == 1,
                                                             ExamResultSummary.deleted_at.is_(None)).all()
            return result

    @staticmethod
    def change_result_stage(result_summary_uid: str, stage: str) -> bool:
        # check stage validit

        with session_scope() as session:

            process = session.query(Process).filter(Process.process_unique_uid == result_summary_uid).first()
            state = session.query(State).filter(State.label == stage).first()

            if can_progress(session, process, state):
                # update result summary exam status
                session.query(ExamResultSummary).filter(ExamResultSummary.uid == result_summary_uid).update(
                    {"exam_status": stage})
                # create process if it does not exist
                if process is None:
                    work_flow = session.query(Workflow).filter(Workflow.name == 'EXAM_FORWARDING').first()
                    process = Process(
                        process_unique_uid=result_summary_uid,
                        workflow=work_flow,
                        description='EXAM_FORWARDING'
                    )
                    session.add(process)
                    session.commit()
                # crate process flow progress
                process_flow = ProcessFlow(
                    state=state,
                    process=process
                )
                session.add(process_flow)
                session.commit()
                return True
            else:
                return False

    def change_program_course_result_stage(self, program_course_id: str, stage: str):
        with session_scope() as session:
            result_summaries = session.query(ExamResultSummary).filter(
                ExamResultSummary.program_course_id == program_course_id).all()
            for result_summary in result_summaries:
                self.change_result_stage(result_summary.uid, stage)
        return True


ExamResultSummaryCrud = ExamResultSummaryService(ExamResultSummary)
