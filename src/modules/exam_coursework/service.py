import base64
import json
from io import BytesIO
from typing import List, Type

import openpyxl
import requests
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill, Protection
from openpyxl.workbook import Workbook
from sqlalchemy import and_
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import aliased
from sqlalchemy import func
from openpyxl.drawing.image import Image

from src.api_routes.program_api import reformat_name
from src.core.config import settings
from src.db.session import session_scope
from src.helpers.utils import to_superscript
from src.models import AcademicYear, ExamCategory, ProgramCourse, ProgramSemester, Course, ExamResultSummary, \
    ProgramCourseAssessment, ExamResult
from src.models.exam_coursework import ExamCoursework
from src.modules.program_semester.service import ProgramSemesterCrud
from src.shared.response_code import ResponseCode
from src.types import StudentCourseWorkOutput, ExamCourseWorkNode, Score, CourseWorkTypeOutput, \
    ExamCourseWorkSearchCriteria, ExcelFile
from src.shared.response import Response


class ExamCourseworkService:

    @staticmethod
    def get_exam_course_work_results(search_criteria: ExamCourseWorkSearchCriteria) -> List[ExamCourseWorkNode]:
        with (session_scope() as session):
            # query = (
            #     session.query(
            #         ExamResultSummary
            #     )
            #     .filter(
            #         ExamResultSummary.program_course_id == search_criteria.program_course_id,
            #         ExamResultSummary.deleted_at.is_(None),
            #     )
            # )
            #
            # if search_criteria.student_uid:
            #     query = query.filter(ExamResultSummary.student_uid == search_criteria.student_uid)
            #
            # if search_criteria.program_course_id:
            #     query = query.filter(ExamResultSummary.program_course_id == search_criteria.program_course_id)
            #
            # results = query.order_by(ExamResultSummary.registration_number.asc()).all()
            #
            # # Retrieve the relevant ExamCoursework records based on the filtered ExamResultSummary records
            # all_results = []
            #
            # for exam_result_summary in results:
            #     exam_coursework_records = (
            #         session.query(ExamCoursework)
            #         .filter_by(student_uid=exam_result_summary.student_uid,
            #                    program_course_id=exam_result_summary.program_course_id)
            #         .all()
            #     )
            #     all_results.append((exam_result_summary, exam_coursework_records))
            # # print("all_results:", all_results)
            # return all_results
            return []

    @staticmethod
    def get_semester_course_results(program_course_uid, info) -> Response[ExcelFile]:
        with (session_scope() as session):
            workbook = Workbook()
            if info.context.user is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Session Expired, refresh this page",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            headers = {
                "Content-Type": "application/json"
            }
            data_obj = {
                "user_uid": str(info.context.user.uid),
                "document_type": "signature"
            }
            payload = json.dumps(data_obj)
            # Send the POST request
            response = requests.post(
                settings.UAA_URi + f'/user/document?user_uid={info.context.user.uid}&document_type=signature',
                headers=headers)

            response.raise_for_status()
            if response.status_code != 200:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Failed To Get User Signature",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            signature_data = response.json()
            if not signature_data['status']:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Please upload your signature before downloading Semester results",
                    data=ExcelFile(base64_data=[], file_name=""),
                )
            # Get Program Semester Information
            program_course = session.query(ProgramCourse.id, ProgramCourse.program_semester_id, ProgramCourse.credit,
                                           ProgramCourse.course_id).filter(
                ProgramCourse.uid == program_course_uid, ProgramCourse.deleted_at.is_(None)).first()
            if program_course is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message="Selected Program Course Is Not Defined",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            theory_coursework_headers_list = []
            ue_headers_list = []
            final_col_to_be_merged = 0
            theory_col_to_be_merged = 0
            ue_col_to_be_merged = 0
            # Get Course Works Type
            course_work_result_type = session.query(ExamCategory.code, ExamCategory.id). \
                join(ProgramCourseAssessment, ProgramCourseAssessment.exam_category_id == ExamCategory.id). \
                filter(ProgramCourseAssessment.program_course_id == program_course.id,
                       ExamCategory.is_ue.is_(False),
                       ProgramCourseAssessment.deleted_at.is_(None), ExamCategory.deleted_at.is_(None)). \
                group_by(ExamCategory.code, ExamCategory.id).all()

            ue_result_type = session.query(ExamCategory.code, ExamCategory.id). \
                join(ExamResult, ExamResult.exam_category_id == ExamCategory.id). \
                filter(ExamResult.program_course_id == program_course.id, ExamResult.number_of_sitting == 1,
                       ExamResult.deleted_at.is_(None), ExamCategory.deleted_at.is_(None)).group_by(
                ExamCategory.code, ExamCategory.id).all()
            # print(ue_result_type)
            theory_over_all_marks = 0
            ue_over_all_marks = 0
            if course_work_result_type:
                for type_code in course_work_result_type:
                    theory_over_all_marks += session.query(func.sum(ProgramCourseAssessment.maximum_score)). \
                                                 filter(ProgramCourseAssessment.program_course_id == program_course.id,
                                                        ProgramCourseAssessment.deleted_at.is_(None),
                                                        ProgramCourseAssessment.exam_category_id == type_code.id).scalar() or 0

                    ass = session.query(ExamCoursework.assessment_number,
                                        func.max(ExamCoursework.weight).label("weight")).filter(
                        ExamCoursework.program_course_id == program_course.id,
                        ExamCoursework.deleted_at.is_(None),
                        ExamCoursework.exam_category_id == type_code.id). \
                        order_by(ExamCoursework.assessment_number.asc()). \
                        group_by(ExamCoursework.assessment_number).all()
                    theory_col_to_be_merged += len(ass) if ass else 0
                    code_info = {'code': type_code.code, 'id': type_code['id'], 'type': 'theory',
                                 'assessment_number': [{'sn': f"{type_code.code}{num.assessment_number}",
                                                        'number': num.assessment_number,
                                                        'weight': int(num.weight)} for num in ass]}
                    theory_coursework_headers_list.append(code_info)

            ue_col_to_be_merged += len(ue_result_type) if ass else 0
            if ue_result_type:
                for type_code in ue_result_type:
                    ue_over_all_marks += session.query(func.sum(ProgramCourseAssessment.maximum_score)). \
                                             filter(ProgramCourseAssessment.program_course_id == program_course.id,
                                                    ProgramCourseAssessment.deleted_at.is_(None),
                                                    ProgramCourseAssessment.exam_category_id == type_code.id).scalar() or 0

                    code_info = {'code': type_code.code, 'id': type_code['id'], 'type': 'ue',
                                 'assessment_number': [{'sn': f"{type_code.code} - 1", 'weight': 1}]}
                    ue_headers_list.append(code_info)

            # print(ue_headers_list)
            results = session.query(ExamResultSummary.student_uid, func.max(ExamResultSummary.gender).label('sex'),
                                    func.max(ExamResultSummary.grade).label('grade'),
                                    func.max(ExamResultSummary.grade_remark).label('grade_remark'),
                                    func.max(ExamResultSummary.cw_score).label('cw_theory'),
                                    func.max(ExamResultSummary.ue_theory).label('ue_theory'),
                                    func.max(ExamResultSummary.grade_point).label('grade_point'),
                                    func.max(ExamResultSummary.total_score).label('total_score'),
                                    func.max(ExamResultSummary.registration_number).label('registration_number'),
                                    func.concat(func.max(ExamResultSummary.first_name), ' ',
                                                func.max(ExamResultSummary.middle_name), ' ',
                                                func.max(ExamResultSummary.last_name)).label('full_name')).filter(
                ExamResultSummary.program_course_id == program_course.id, ExamResultSummary.deleted_at.is_(None)). \
                group_by(ExamResultSummary.student_uid).order_by(func.max(ExamResultSummary.first_name).asc()).all()

            worksheet = workbook.active
            # Set the orientation to landscape
            worksheet.print_options.horizontalCentered = True
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

            # Set the font style to Times New Roman
            font = Font(name="Times New Roman", size=12)
            small_font = Font(name="Times New Roman", size=10)
            fill_color = PatternFill(start_color='FF999999', end_color='FF999999', fill_type='solid')
            font_border = Font(name="Times New Roman", bold=True, size=12)
            small_font_border = Font(name="Times New Roman", bold=True, size=10)
            # Set the border style
            border = Border(left=Side(border_style="thin"),
                            right=Side(border_style="thin"),
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))

            # Merge cells from C1 to J1
            worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
            # Merge cells from C2 to J2
            worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=10)

            merged_cell_h1 = worksheet.cell(row=1, column=3, value="SOKOINE UNIVERSITY OF AGRICULTURE")
            merged_cell_h1.alignment = Alignment(horizontal='center')
            merged_cell_h1.font = font
            # Merge cells from C2 to J2
            merged_cell_h2 = worksheet.cell(row=2, column=3, value="P.O. BOX 3000, CHUO KIKUU, MOROGORO, TANZANIA")
            merged_cell_h2.alignment = Alignment(horizontal='center')
            merged_cell_h2.font = font
            count_rows = 4

            # Define the vertical headers
            vertical_headers = ["Course Results", "Program", "Course", "Academic Year", "Study Year", "Semester",
                                "Course Credit"]

            program_semester = session.query(ProgramSemester).filter(
                ProgramSemester.id == program_course.program_semester_id, ProgramSemester.deleted_at.is_(None)).first()
            course = session.query(Course.name, Course.code).filter(Course.id == program_course.course_id).first()
            file_name = f"{program_semester.program.name}({program_semester.program.code}) {program_semester.academic_year.name} - YEAR {program_semester.study_year} SEMESTER {program_semester.semester} COURSE {course.code}  RESULT"

            if results is None:
                return Response(
                    status=False,
                    code=ResponseCode.FAILURE,
                    message=f"No Result Found For {file_name}",
                    data=ExcelFile(base64_data=[], file_name=""),
                )

            # Sample data for the vertical header
            data = {
                "Course Results": "",
                "Program": str(program_semester.program.name) + " (" + str(program_semester.program.code) + ")",
                "Course": str(course.name) + " (" + str(course.code) + ")",
                "Academic Year": str(program_semester.academic_year.name),
                "Study Year": str(program_semester.study_year),
                "Semester": str(program_semester.semester),
                "Course Credit": str(program_course.credit)
            }
            worksheet.sheet_view.showGridLines = False
            # Generate the data for the vertical header
            vertical_data = [data[header] for header in vertical_headers]
            for row, header in enumerate(vertical_headers, start=4):
                count_rows += 1
                cell = worksheet.cell(row=row, column=3)
                cell.value = header
                cell.alignment = Alignment(horizontal='left')
                if count_rows == 5:
                    cell.font = font_border
                else:
                    cell.font = font
                cell.border = None
                cell.protection = Protection(locked=False)
            for row, value in enumerate(vertical_data, start=4):
                cell = worksheet[f"F{row}"]
                cell.value = value
                cell.font = font
                cell.border = None
                cell.protection = Protection(locked=False)

            # SN
            worksheet.merge_cells(start_row=count_rows, start_column=1, end_row=count_rows + 2,
                                  end_column=1)
            summary_text = worksheet.cell(row=count_rows, column=1, value="SNo")
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = small_font_border
            summary_text.border = border
            # Name
            worksheet.merge_cells(start_row=count_rows, start_column=2, end_row=count_rows + 2,
                                  end_column=2)
            text = worksheet.cell(row=count_rows, column=2, value="Name")
            text.alignment = Alignment(horizontal='center')
            text.font = small_font_border
            text.border = border

            # SN
            worksheet.merge_cells(start_row=count_rows, start_column=3, end_row=count_rows + 2,
                                  end_column=3)
            text = worksheet.cell(row=count_rows, column=3, value="Reg. No.")
            text.alignment = Alignment(horizontal='center')
            text.font = small_font_border
            text.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=4, end_row=count_rows + 2,
                                  end_column=4)
            text = worksheet.cell(row=count_rows, column=4, value="Sex")
            text.alignment = Alignment(horizontal='center')
            text.font = small_font_border
            text.border = border

            # Theory Course Work
            theory_col_to_be_merged += len(theory_coursework_headers_list) + 1
            worksheet.merge_cells(start_row=count_rows, start_column=5, end_row=count_rows,
                                  end_column=theory_col_to_be_merged + 4)
            text = worksheet.cell(row=count_rows, column=5, value="Theory Course Work")
            text.alignment = Alignment(horizontal='center')
            text.font = small_font_border
            text.border = border

            worksheet.merge_cells(start_row=count_rows, start_column=(theory_col_to_be_merged + 4) + 1,
                                  end_row=count_rows,
                                  end_column=theory_col_to_be_merged + 5 + ue_col_to_be_merged + 1)
            text = worksheet.cell(row=count_rows, column=(theory_col_to_be_merged + 4) + 1, value="UE Theory")
            text.alignment = Alignment(horizontal='center')
            text.font = small_font_border
            text.border = border

            nex_row = count_rows
            ue_start_column = 0
            if theory_coursework_headers_list:
                header_column = 4
                row_no = 0
                nex_row += 1
                col_no = 0
                for header in theory_coursework_headers_list:
                    row_no += 1
                    total_weight = 0
                    for sn in header['assessment_number']:
                        col_no += 1
                        text = worksheet.cell(row=nex_row, column=header_column + col_no, value=sn['sn'])
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font_border
                        text.border = border
                        total_weight += sn['weight']
                        text = worksheet.cell(row=nex_row + 1, column=header_column + col_no, value=sn['weight'])
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font_border
                        text.border = border

                        # print(header['assessment_number'])
                    col_no += 1
                    text = worksheet.cell(row=nex_row, column=header_column + col_no, value="TOTAL")
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font_border
                    text.border = border
                    # #
                    text = worksheet.cell(row=nex_row + 1, column=header_column + col_no, value=total_weight)
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font_border
                    text.border = border

                ue_start_column = (header_column + col_no) + 1

                start_row = nex_row
                start_column = (header_column + col_no) + 1
                # print(f"{start_row}:{start_column} - {start_row+1}:{start_column}")

                worksheet.merge_cells(start_row=start_row, start_column=start_column,
                                      end_row=nex_row + 1,
                                      end_column=start_column)

                text = worksheet.cell(row=nex_row, column=(header_column + col_no) + 1,
                                      value=f"Marks (x/{theory_over_all_marks})")
                text.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                text.font = small_font_border
                text.border = border

                worksheet.column_dimensions[
                    f"{openpyxl.utils.get_column_letter((header_column + col_no) + 1)}"].width = 5

            # ue_start_column +=
            nex_row = count_rows
            if ue_headers_list:
                header_column = ue_start_column
                row_no = 0
                nex_row += 1
                col_no = 0
                for header in ue_headers_list:
                    row_no += 1
                    total_weight = 0
                    for sn in header['assessment_number']:
                        col_no += 1
                        text = worksheet.cell(row=nex_row, column=header_column + col_no, value=sn['sn'])
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font_border
                        text.border = border
                        total_weight += sn['weight']
                        text = worksheet.cell(row=nex_row + 1, column=header_column + col_no, value=sn['weight'])
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font_border
                        text.border = border

                        # print(header['assessment_number'])
                    col_no += 1
                    text = worksheet.cell(row=nex_row, column=header_column + col_no, value="TOTAL")
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font_border
                    text.border = border
                    worksheet.column_dimensions[
                        f"{openpyxl.utils.get_column_letter((header_column + col_no))}"].width = 6
                    # #
                    text = worksheet.cell(row=nex_row + 1, column=header_column + col_no, value=total_weight)
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font_border
                    text.border = border

                start_row = nex_row
                start_column = (header_column + col_no) + 1
                # print(f"{start_row}:{start_column} - {start_row+1}:{start_column}")

                worksheet.merge_cells(start_row=start_row, start_column=start_column,
                                      end_row=nex_row + 1,
                                      end_column=start_column)
                text = worksheet.cell(row=nex_row, column=(header_column + col_no) + 1,
                                      value=f"Marks (x/{ue_over_all_marks})")
                text.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                text.font = small_font_border
                text.border = border
                worksheet.column_dimensions[
                    f"{openpyxl.utils.get_column_letter((header_column + col_no) + 1)}"].width = 5

                worksheet.column_dimensions[
                    f"{openpyxl.utils.get_column_letter((header_column + col_no) + 2)}"].width = 5

                final_list = ['Total', 'Grade', 'Grade Point', 'Remarks']
                sn = 1
                for ln in final_list:
                    sn += 1
                    worksheet.merge_cells(start_row=start_row, start_column=(header_column + col_no) + sn,
                                          end_row=nex_row + 1,
                                          end_column=(header_column + col_no) + sn)
                    text = worksheet.cell(row=start_row, column=(header_column + col_no) + sn,
                                          value=ln)
                    text.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    text.font = small_font_border
                    text.border = border

            count_rows += 2
            count = 0
            count_rows += 1
            grade_a_male = 0
            grade_a_female = 0
            grade_b_male = 0
            grade_b_female = 0
            grade_bp_male = 0
            grade_bp_female = 0
            grade_c_male = 0
            grade_c_female = 0
            grade_fail_pt_male = 0
            grade_fail_pt_female = 0
            grade_d_male = 0
            grade_d_female = 0
            grade_e_male = 0
            grade_e_female = 0
            incomplete_male = 0
            incomplete_female = 0
            for row, item in enumerate(results, start=count_rows):
                count += 1
                # print(item)
                courses_under_probation = ""
                total_credit_hrs_taken = 0
                total_credit_hrs_acquired = 0
                total_failed_core_subject = 0
                failed_subjects = 0
                passed_subjects = 0
                remark_status = 0
                # count_rows += 1
                worksheet[f"A{row}"] = count
                worksheet[f"B{row}"] = reformat_name(item['full_name'])
                worksheet[f"C{row}"] = item['registration_number']
                worksheet[f"D{row}"] = item['sex'][0]
                sex = item['sex'][0].upper()
                if sex == 'M':
                    if item['grade'].upper() == 'A':
                        grade_a_male += 1
                    if item['grade'].upper() == 'B':
                        grade_b_male += 1
                    if item['grade'].upper() == 'B+':
                        grade_bp_male += 1
                    if item['grade'].upper() == 'C':
                        grade_c_male += 1
                    if item['grade'].upper() == 'D':
                        grade_d_male += 1
                    if item['grade'].upper() == 'E':
                        grade_e_male += 1
                    if item['grade'].upper() == 'I':
                        incomplete_male += 1
                else:
                    if item['grade'].upper() == 'A':
                        grade_a_female += 1
                    if item['grade'].upper() == 'B':
                        grade_b_female += 1
                    if item['grade'].upper() == 'B+':
                        grade_bp_female += 1
                    if item['grade'].upper() == 'C':
                        grade_c_female += 1
                    if item['grade'].upper() == 'D':
                        grade_d_female += 1
                    if item['grade'].upper() == 'E':
                        grade_e_female += 1
                    if item['grade'].upper() == 'I':
                        incomplete_female += 1

                for col in range(1, 5):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = small_font
                    cell.border = border
                col = 4
                count_rows += 1
                col_no = 4
                for header in theory_coursework_headers_list:
                    row_no += 1
                    total_score = 0
                    no_of_scores = 0
                    for ass in header['assessment_number']:
                        col_no += 1
                        value = '-'
                        # Check Scores
                        if header['type'] == 'theory':
                            ass_score = session.query(ExamCoursework.score).filter(
                                ExamCoursework.program_course_id == program_course.id,
                                ExamCoursework.student_uid == item['student_uid'],
                                ExamCoursework.assessment_number == ass['number'],
                                ExamCoursework.exam_category_id == header['id']).first()
                            if ass_score:
                                value = ass_score.score
                                no_of_scores += 1
                                total_score += ass_score.score
                            # print(f"{item['registration_number']} - {ass['sn']} - {ass_score}")
                        text = worksheet.cell(row=row, column=col_no, value=value)
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font
                        text.border = border
                        worksheet.column_dimensions[
                            f"{openpyxl.utils.get_column_letter(col_no)}"].width = 5
                    col_no += 1
                    avg = 0
                    if no_of_scores > 0:
                        avg = total_score / no_of_scores
                    text = worksheet.cell(row=row, column=col_no, value=round(avg, 2))
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font
                    text.border = border
                    worksheet.column_dimensions[
                        f"{openpyxl.utils.get_column_letter(col_no)}"].width = 6

                    # exam_summary = session.query(ExamResultSummary.cw_theory).filter(
                    #     ExamResultSummary.program_course_id == program_course.id,
                    #     ExamResultSummary.student_uid == item['student_uid']).first()
                    score = '-'
                    if item['cw_theory']:
                        score = round(item['cw_theory'], 2)
                    text = worksheet.cell(row=row, column=col_no + 1, value=score)
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font
                    text.border = border

                col_no = ue_start_column
                for header in ue_headers_list:
                    # row_no += 1
                    total_score = 0
                    no_of_scores = 0
                    for ass in header['assessment_number']:
                        col_no += 1
                        value = '-'
                        # Check Scores
                        if header['type'] == 'ue':
                            ue_score = session.query(ExamResult.score).filter(
                                ExamResult.program_course_id == program_course.id,
                                ExamResult.student_uid == item['student_uid'],
                                ExamResult.number_of_sitting == 1,
                                ExamResult.exam_category_id == header['id']).first()
                            if ue_score:
                                value = ue_score.score
                                no_of_scores += 1
                                total_score += ass_score.score
                        text = worksheet.cell(row=row, column=col_no, value=value)
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font
                        text.border = border
                        worksheet.column_dimensions[
                            f"{openpyxl.utils.get_column_letter(col_no)}"].width = 5

                    col_no += 1
                    avg = 0
                    if no_of_scores > 0:
                        avg = round(total_score / no_of_scores, 2)
                    text = worksheet.cell(row=row, column=col_no, value=avg)
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font
                    text.border = border

                    score = '-'
                    if item['ue_theory']:
                        score = round(item['ue_theory'], 2)
                    text = worksheet.cell(row=row, column=col_no + 1, value=score)
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font
                    text.border = border
                    grade = item['grade']
                    if '|' in grade:
                        pipe_index = grade.index('|')
                        e_value = grade[:pipe_index]
                        pt_value = grade[pipe_index + 1:]
                        pt_value = pt_value.replace('|', '')
                        combined_value = f"{e_value}{to_superscript(pt_value)}"
                        # Set the combined value to the cell
                        grade = combined_value

                    total_list_to_display = [grade, item['grade_point'], item['grade_remark']]
                    total_score = '-'
                    if item['total_score']:
                        total_score = round(item['total_score'], 2)
                    text = worksheet.cell(row=row, column=col_no + 2, value=total_score)
                    text.alignment = Alignment(horizontal='center')
                    text.font = small_font
                    text.border = border
                    no = 2
                    for ln in total_list_to_display:
                        no += 1
                        text = worksheet.cell(row=row, column=col_no + no, value=ln)
                        text.alignment = Alignment(horizontal='center')
                        text.font = small_font
                        text.border = border

            # print(theory_over_all_marks)

            worksheet.column_dimensions['A'].width = 4
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 13
            worksheet.column_dimensions['D'].width = 5
            # worksheet.column_dimensions['F'].width = 11

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.protection = Protection(locked=True)

                # Protect the worksheet to make cells not editable
                worksheet.protection.sheet = True

            count_rows += 1
            worksheet.merge_cells(start_row=count_rows, start_column=1,
                                  end_row=count_rows,
                                  end_column=8)
            text = worksheet.cell(row=count_rows, column=1, value="Summary")
            text.alignment = Alignment(horizontal='center', vertical='center')
            text.font = small_font_border
            text.border = border

            count_rows += 1
            summary_headers = ['Grade', 'Male', 'Female', '%', 'Remarks', 'Count', '%']

            sn = 0
            st_col = 0
            for sh in summary_headers:
                sn += 1
                st_col += 1
                if sn == 1:
                    worksheet.merge_cells(start_row=count_rows, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col + 1)

                text = worksheet.cell(row=count_rows, column=st_col, value=sh)
                text.alignment = Alignment(horizontal='center', vertical='center')
                text.font = small_font_border
                text.border = border

                if sn == 1:
                    st_col += 1

            total_male = (
                    grade_a_male + grade_bp_male + grade_b_male + grade_c_male + grade_fail_pt_male + grade_d_male + grade_e_male + incomplete_male)
            total_female = (
                    grade_a_female + grade_bp_female + grade_b_female + grade_c_female + grade_fail_pt_female + grade_d_female + grade_e_female + incomplete_female)

            grade_data = [{'grade': 'A', 'male': grade_a_male, 'female': grade_a_female},
                          {'grade': 'B+', 'male': grade_bp_male, 'female': grade_bp_female},
                          {'grade': 'B', 'male': grade_b_male, 'female': grade_b_female},
                          {'grade': 'C', 'male': grade_c_male, 'female': grade_c_female},
                          {'grade': 'Fail (P|T)', 'male': grade_fail_pt_male, 'female': grade_fail_pt_female},
                          {'grade': 'D', 'male': grade_d_male, 'female': grade_d_female},
                          {'grade': 'E', 'male': grade_e_male, 'female': grade_e_female},
                          {'grade': 'Incomplete', 'male': incomplete_male, 'female': incomplete_female},
                          {'grade': 'Total', 'male': total_male, 'female': total_female}, ]

            rows = 0
            total_fail = 0
            total = 0
            total_pec = 0
            total_all = 0

            for sh in grade_data:
                rows += 1
                count_rows += 1
                st_col = 1
                total += sh['male'] + sh['female']
                worksheet.merge_cells(start_row=count_rows, start_column=st_col,
                                      end_row=count_rows,
                                      end_column=st_col + 1)
                text = worksheet.cell(row=count_rows, column=st_col, value=sh['grade'])
                text.alignment = Alignment(horizontal='left')
                text.font = small_font
                text.border = border

                st_col += 2
                text = worksheet.cell(row=count_rows, column=st_col, value=sh['male'])
                text.alignment = Alignment(horizontal='center')
                text.font = small_font_border
                text.border = border
                st_col += 1
                text = worksheet.cell(row=count_rows, column=st_col, value=sh['female'])
                text.alignment = Alignment(horizontal='center')
                text.font = small_font_border
                text.border = border

                # print('len(result)', len(results))
                percent = ((sh['male'] + sh['female']) / len(results)) * 100
                st_col += 1
                text = worksheet.cell(row=count_rows, column=st_col, value=round(percent, 2))
                text.alignment = Alignment(horizontal='center')
                text.font = small_font_border
                text.border = border
                st_col += 1
                if rows == 4:
                    worksheet.merge_cells(start_row=count_rows - 3, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col)
                    text = worksheet.cell(row=count_rows - 3, column=st_col, value="PASS")
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1
                    worksheet.merge_cells(start_row=count_rows - 3, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col)
                    text = worksheet.cell(row=count_rows - 3, column=st_col, value=total)
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border

                    st_col += 1
                    worksheet.merge_cells(start_row=count_rows - 3, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col)
                    # print('len(result)-2', len(results))
                    percent = (total / len(results)) * 100
                    total_pec += percent
                    text = worksheet.cell(row=count_rows - 3, column=st_col, value=round(percent, 2))
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    total_all += total
                    total = 0

                if rows == 7:
                    worksheet.merge_cells(start_row=count_rows - 2, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col)
                    text = worksheet.cell(row=count_rows - 2, column=st_col, value="FAIL")
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1
                    worksheet.merge_cells(start_row=count_rows - 2, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col)
                    text = worksheet.cell(row=count_rows - 2, column=st_col, value=total)
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1
                    worksheet.merge_cells(start_row=count_rows - 2, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col)
                    # print('len(result)-3', len(results))

                    percent = (total / len(results)) * 100
                    total_pec += percent
                    text = worksheet.cell(row=count_rows - 2, column=st_col, value=round(percent, 2))
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    total_all += total
                    total = 0

                if rows == 8:
                    text = worksheet.cell(row=count_rows, column=st_col, value="INCOMPLETE")
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1
                    text = worksheet.cell(row=count_rows, column=st_col, value=total)
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1

                    # print('len(result) - 4', len(results))
                    percent = (total / len(results)) * 100
                    total_pec += percent
                    text = worksheet.cell(row=count_rows, column=st_col, value=round(percent, 2))
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    total_all += total
                    total = 0
                #
                if rows == 9:
                    text = worksheet.cell(row=count_rows, column=st_col, value="")
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1
                    text = worksheet.cell(row=count_rows, column=st_col, value=total_all)
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 1
                    text = worksheet.cell(row=count_rows, column=st_col, value=total_pec)
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border

            count_rows += 3
            signature_row = count_rows
            worksheet.merge_cells(start_row=count_rows, start_column=1,
                                  end_row=count_rows,
                                  end_column=8)
            text = worksheet.cell(row=count_rows, column=1, value="Key")
            text.alignment = Alignment(horizontal='center', vertical='center')
            text.font = small_font_border
            text.border = border
            key_data = [{'name': 'Code', 'dec': 'Description'},
                        {'name': 'ASS', 'dec': 'Assignment'},
                        {'name': 'TT', 'dec': 'Theory Test'},
                        {'name': 'UE', 'dec': 'UE Theory'},
                        {'name': 'SR', 'dec': 'Seminar Report'}]
            sn = 0
            for key in key_data:
                count_rows += 1
                sn += 1
                st_col = 0
                st_col += 1
                if sn == 1:
                    worksheet.merge_cells(start_row=count_rows, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col + 1)
                    text = worksheet.cell(row=count_rows, column=st_col, value=key['name'])
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                    st_col += 2
                    worksheet.merge_cells(start_row=count_rows, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col + 5)
                    text = worksheet.cell(row=count_rows, column=st_col, value=key['dec'])
                    text.alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
                    text.font = small_font_border
                    text.border = border
                else:
                    worksheet.merge_cells(start_row=count_rows, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col + 1)
                    text = worksheet.cell(row=count_rows, column=st_col, value=key['name'])
                    text.alignment = Alignment(horizontal='left')
                    text.font = small_font
                    text.border = border
                    st_col += 2
                    worksheet.merge_cells(start_row=count_rows, start_column=st_col,
                                          end_row=count_rows,
                                          end_column=st_col + 5)
                    text = worksheet.cell(row=count_rows, column=st_col, value=key['dec'])
                    text.alignment = Alignment(horizontal='left')
                    text.font = small_font
                    text.border = border

            column_no = 14
            # Decode the Base64 image
            base64_image = signature_data['data']['base64doc']
            decoded_image = base64.b64decode(base64_image)

            # Load the image into BytesIO
            image_stream = BytesIO(decoded_image)

            # Create an Image object from BytesIO
            img = Image(image_stream)

            # Calculate the coordinates of the bottom center cell
            bottom_center_cell = openpyxl.utils.coordinate_to_tuple(
                f'{openpyxl.utils.get_column_letter(column_no)}{signature_row + 1}')

            # Calculate the column letter for the anchor cell (e.g., "B" for column number 2)
            anchor_column = openpyxl.utils.get_column_letter(14)

            # Set the row number for the anchor cell
            anchor_row = bottom_center_cell[0]

            # Set the anchor cell for the image to the bottom center cell
            anchor_cell = f'{anchor_column}{anchor_row}'
            # Add image to the worksheet and set the anchor cell
            worksheet.add_image(img, anchor_cell)
            worksheet[anchor_cell].alignment = Alignment(horizontal='center', vertical='bottom')

            name_row = signature_row + 4
            colum_no = 16
            worksheet.merge_cells(start_row=name_row, start_column=10, end_row=name_row, end_column=colum_no + 2)
            summary_text = worksheet.cell(row=name_row, column=10, value=info.context.user.full_name)
            summary_text.alignment = Alignment(horizontal='center')
            summary_text.font = font_border

            # pr
            # print(count_rows)
            # Save the workbook to a BytesIO buffer
            file_buffer = BytesIO()
            workbook.save(file_buffer)
            file_buffer.seek(0)

            # Convert the Excel file to Base64 string
            file_data = file_buffer.getvalue()
            base64_data = base64.b64encode(file_data).decode()

            return Response(
                status=True,
                code=ResponseCode.SUCCESS,
                message="Semester Exam Results Retrieved Successfully",
                data=ExcelFile(base64_data=base64_data, file_name=file_name.upper())
            )

            # print(program_course.program_course_assessments)
            # if not program_semester:
            return Response(
                status=False,
                code=ResponseCode.FAILURE,
                message="Program Semester Is Not Defined",
                data=ExcelFile(base64_data=[], file_name=""),
            )

    @staticmethod
    def get_student_exam_course_work_results(student_uid) -> List[ExamCoursework]:
        with session_scope() as session:
            result = session.query(ExamCoursework).filter(ExamCoursework.student_uid == student_uid,
                                                          ExamCoursework.deleted_at.is_(None)).all()
            return result

    @staticmethod
    def get_student_active_semester_course_work_results(input) -> List[StudentCourseWorkOutput]:
        with session_scope() as session:
            academic_year = session.query(AcademicYear.id).filter(AcademicYear.uid == input.academic_year_uid).first()
            if academic_year is None:
                return Response(
                    status=False,
                    code=ResponseCode.NO_RECORD_FOUND,
                    message="Academic year not found",
                    data=[]
                )
            # print(current_academic_year)
            result = (
                session.query(
                    ProgramCourse.id.label("program_course_id"),
                    Course.code.label("course_code"),
                    Course.name.label("course_name"),
                )
                .join(Course, ProgramCourse.course_id == Course.id)
                .join(ExamCoursework, ExamCoursework.program_course_id == ProgramCourse.id)
                .join(ExamCategory, ExamCoursework.exam_category_id == ExamCategory.id)
                .join(ProgramSemester, ProgramSemester.id == ProgramCourse.program_semester_id)
                .filter(ExamCoursework.student_uid == input.student_uid,
                        ProgramCourse.deleted_at.is_(None),
                        ExamCategory.is_ue.is_(False),
                        ProgramSemester.semester == input.semester,
                        ProgramSemester.study_year == input.study_year,
                        ProgramSemester.academic_year_id == academic_year.id)
                .group_by(ProgramCourse.id, Course.code, Course.name)
                .all()
            )

            course_data = []
            if result:
                for course in result:
                    course_type = session.query(ExamCategory.name, ExamCategory.id). \
                        join(ExamCoursework, ExamCategory.id == ExamCoursework.exam_category_id). \
                        filter(
                        and_(
                            ExamCoursework.program_course_id == course.program_course_id,
                            ExamCoursework.student_uid == input.student_uid,
                            ExamCategory.is_ue.is_(False),
                        )
                    ).group_by(ExamCategory.name, ExamCategory.id).all()
                    # print(course_type)
                    course_type_data = []
                    if course_type:
                        for cType in course_type:
                            scores = session.query(ExamCoursework.score, ExamCoursework.overall_marks,
                                                   ExamCoursework.assessment_number). \
                                filter(ExamCoursework.student_uid == input.student_uid,
                                       ExamCoursework.exam_category_id == cType.id,
                                       ExamCoursework.program_course_id == course.program_course_id). \
                                order_by(ExamCoursework.assessment_number.asc()).all()
                            if scores:
                                scoreData = []
                                for mark in scores:
                                    s = {
                                        "score": mark.score,
                                        "overall_marks": mark.overall_marks,
                                        "assessment_number": mark.assessment_number
                                    }
                                    scoreData.append(s)
                                cInfo = {
                                    "type": cType.name,
                                    "score": scoreData
                                }
                                course_type_data.append(cInfo)

                        c = {
                            "course_code": course.course_code,
                            "course_name": course.course_name,
                            "course_work_type": course_type_data
                        }
                        course_data.append(c)
                student_course_work_output_list = []
                for item in course_data:
                    course_work_type_list = []
                    for work_type in item['course_work_type']:
                        score_list = [Score(**score) for score in work_type['score']]
                        course_work_type_list.append(CourseWorkTypeOutput(type=work_type['type'], scores=score_list))
                    student_course_work_output_list.append(StudentCourseWorkOutput(course_code=item['course_code'],
                                                                                   course_name=item['course_name'],
                                                                                   course_work_type=course_work_type_list))
                # print(course_data)
                # return student_course_work_output_list
                return Response(
                    status=True,
                    code=ResponseCode.SUCCESS,
                    message="Course Works Results Retrieved Successfully",
                    data=student_course_work_output_list
                )

            return Response(
                status=False,
                code=ResponseCode.NO_RECORD_FOUND,
                message="Course work has not been uploaded yet.",
                data=[]
            )


exam_coursework_service = ExamCourseworkService()
